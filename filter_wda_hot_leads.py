import csv
import os
import re
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Set console output encoding to UTF-8
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_CSV = os.path.join(SCRIPT_DIR, "wda_employers.csv")
OUTPUT_CSV = os.path.join(SCRIPT_DIR, "wda_hot_leads.csv")
OUTPUT_XLSX = os.path.join(SCRIPT_DIR, "wda_hot_leads.xlsx")

def score_lead(row):
    score = 0
    reasons = []
    
    # 1. Preferred Nationality (Quốc tịch mong muốn)
    nationality = row.get("Quốc tịch mong muốn", "").strip()
    if "越南" in nationality:
        score += 30
        reasons.append("Ưu tiên Việt Nam (+30)")
    elif "無限制" in nationality or "无限制" in nationality:
        score += 10
        reasons.append("Không giới hạn quốc tịch (+10)")
        
    # 2. Succession Quota (Số lượng tuyển / Quota)
    quota_str = row.get("Số lượng tuyển (Quota)", "").strip()
    quota = 0
    match = re.search(r'\d+', quota_str)
    if match:
        quota = int(match.group())
        
    if quota >= 50:
        score += 25
        reasons.append(f"Số lượng tuyển lớn {quota} người (+25)")
    elif quota >= 20:
        score += 15
        reasons.append(f"Số lượng tuyển vừa {quota} người (+15)")
        
    # 3. Email contact (Email liên hệ)
    email = row.get("Email liên hệ", "").strip()
    if email and "@" in email:
        score += 10
        reasons.append("Có Email HR liên hệ (+10)")
        
    # 4. No Agency / Direct Contact (Không thấy môi giới)
    agency_phone = row.get("SĐT môi giới", "").strip()
    if not agency_phone or "不顯示" in agency_phone or agency_phone == "":
        score += 30
        reasons.append("Không qua môi giới / Liên hệ trực tiếp (+30)")
        
    # 5. Job Category (Ngành nghề)
    category = row.get("Ngành nghề", "").strip()
    if "製造工" in category or "sản xuất" in category.lower() or "chế tạo" in category:
        score += 20
        reasons.append("Ngành sản xuất/chế tạo (+20)")
        
    return score, ", ".join(reasons)

def main():
    if not os.path.exists(INPUT_CSV):
        print(f"[-] Không tìm thấy file dữ liệu gốc: {INPUT_CSV}")
        return
        
    print("[*] Đang đọc dữ liệu cào WDA và tiến hành chấm điểm Leads...")
    
    rows = []
    with open(INPUT_CSV, mode='r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        for r in reader:
            rows.append(r)
            
    print(f"[+] Đã tải {len(rows)} bản ghi để chấm điểm.")
    
    scored_rows = []
    for r in rows:
        score, reasons = score_lead(r)
        r["Điểm đánh giá"] = score
        r["Lý do xếp hạng"] = reasons
        
        # Determine lead level
        if score >= 80:
            r["Xếp loại Lead"] = "HOT LEAD 🔥"
        elif score >= 50:
            r["Xếp loại Lead"] = "WARM LEAD ⚡"
        else:
            r["Xếp loại Lead"] = "COLD LEAD ❄️"
            
        scored_rows.append(r)
        
    # Sort rows by score descending
    scored_rows.sort(key=lambda x: x["Điểm đánh giá"], reverse=True)
    
    # Write to CSV
    output_fields = ["Điểm đánh giá", "Xếp loại Lead", "Lý do xếp hạng"] + [f for f in fieldnames if f != "No."]
    # Add index "No."
    for idx, r in enumerate(scored_rows, 1):
        r["No."] = idx
        
    output_fields = ["No."] + output_fields
    
    print(f"[*] Ghi danh sách Leads đã chấm điểm ra CSV: {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, mode='w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=output_fields)
        writer.writeheader()
        for r in scored_rows:
            # Clean up dict keys to match output_fields
            row_data = {k: r.get(k, "") for k in output_fields}
            writer.writerow(row_data)
            
    # Write to Excel with formatting
    print(f"[*] Ghi danh sách Leads đã chấm điểm ra Excel: {OUTPUT_XLSX}...")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Leads Ranked"
    
    sheet.append(output_fields)
    
    # Stylings
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    for col_idx in range(1, len(output_fields) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    hot_fill = PatternFill(start_color="FFD8D8", end_color="FFD8D8", fill_type="solid") # light red
    warm_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # light yellow
    cold_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # light green
    
    for r_idx, r in enumerate(scored_rows, 2):
        row_data = [r.get(k, "") for k in output_fields]
        sheet.append(row_data)
        
        # Color rating cell
        lead_type = r.get("Xếp loại Lead", "")
        rating_cell = sheet.cell(row=r_idx, column=2)
        score_cell = sheet.cell(row=r_idx, column=1)
        
        if "HOT" in lead_type:
            rating_cell.fill = hot_fill
            score_cell.fill = hot_fill
        elif "WARM" in lead_type:
            rating_cell.fill = warm_fill
            score_cell.fill = warm_fill
        else:
            rating_cell.fill = cold_fill
            score_cell.fill = cold_fill
            
    # Save
    wb.save(OUTPUT_XLSX)
    print("[SUCCESS] Đã chấm điểm và xếp hạng danh sách Leads thành công!")

if __name__ == "__main__":
    main()
