import csv
import glob
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Set console output encoding to UTF-8
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

def safe_print(msg):
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        try:
            print(msg.encode(sys.stdout.encoding or 'utf-8', errors='replace').decode(sys.stdout.encoding or 'utf-8'), flush=True)
        except Exception:
            print(msg.encode('ascii', errors='replace').decode('ascii'), flush=True)

def sort_file_by_email(file_path, id_col, email_col):
    if not os.path.exists(file_path):
        safe_print(f"[-] File not found: {file_path}")
        return False
        
    safe_print(f"[*] Sorting: {os.path.basename(file_path).encode('ascii', errors='replace').decode('ascii')}")
    
    with open(file_path, mode='r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)
        
    if email_col not in fieldnames or id_col not in fieldnames:
        safe_print(f"  [-] Column '{email_col}' or '{id_col}' not found in headers!")
        return False
        
    # Split rows
    rows_with_email = []
    rows_without_email = []
    
    for row in rows:
        email = row.get(email_col, "").strip()
        if email:
            rows_with_email.append(row)
        else:
            rows_without_email.append(row)
            
    sorted_rows = rows_with_email + rows_without_email
    
    # Re-index sequence column
    for idx, row in enumerate(sorted_rows, 1):
        row[id_col] = str(idx)
        
    # Save CSV atomically
    temp_path = file_path + ".tmp"
    try:
        with open(temp_path, mode='w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(sorted_rows)
        if os.path.exists(file_path):
            os.remove(file_path)
        os.rename(temp_path, file_path)
        safe_print(f"  [+] Sorted {len(rows)} rows. Pushed {len(rows_with_email)} rows with emails to the top.")
        return True
    except Exception as e:
        safe_print(f"  [-] Failed to write sorted file: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return False

def regenerate_korean_excel():
    csv_path = "korean_agencies.csv"
    xlsx_path = "korean_agencies.xlsx"
    if not os.path.exists(csv_path):
        return
        
    safe_print(f"[*] Regenerating Excel: {xlsx_path}")
    try:
        with open(csv_path, mode="r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            fieldnames = reader.fieldnames
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Naver Map Recruiter"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws.append(fieldnames)
        for col_num, header_name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        for r_idx, row in enumerate(rows, 1):
            row_data = [row.get(fn, "") for fn in fieldnames]
            ws.append(row_data)
            row_num = r_idx + 1
            
            is_even = (r_idx % 2 == 0)
            for col_num in range(1, len(fieldnames) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_all
                if is_even:
                    cell.fill = fill_zebra
                
                if col_num in [1, 4, 8, 9]: # STT, Phone, Location, Keyword
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws.row_dimensions[1].height = 28
        for r in range(2, len(rows) + 2):
            ws.row_dimensions[r].height = 20
            
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(xlsx_path)
        safe_print(f"  [+] Regenerated Excel successfully.")
    except Exception as e:
        safe_print(f"  [-] Failed to write Excel: {e}")

def main():
    # 1. Sort Ireland Recruiter CSV
    files = glob.glob("*(ch*M*gi*ireland*CleanData.csv")
    for f in files:
        sort_file_by_email(f, "No.", "Email")
        
    # 2. Sort Ireland Farm CSV
    files = glob.glob("*(ch*)*n*tr*ireland*CleanData.csv")
    for f in files:
        sort_file_by_email(f, "No.", "Email")
        
    # 3. Sort Korean Agencies CSV
    files = glob.glob("*korean_agencies.csv")
    for f in files:
        if sort_file_by_email(f, "STT / No.", "Website"): # In Korean map, Website column contains URL (which represents matching contact info)
            regenerate_korean_excel()

if __name__ == "__main__":
    main()
