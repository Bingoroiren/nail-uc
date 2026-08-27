import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Domains that belong to news outlets, directories or software agencies
JUNK_DOMAINS = [
    "news2day.co.kr", "newsprime.co.kr", "insight.co.kr", "siminilbo.co.kr", 
    "naeil.com", "allthatcompany.com", "bizwiki.co.kr", "bizwiki.co", "foxsoft.kr",
    "kjuso.kr", "ggilbo.com", "hankooki.com", "cnews.co.kr", "boannews.com"
]

import re
import urllib.parse

SOCIAL_DOMAINS = {
    'facebook.com', 'fb.com', 'fb.me',
    'instagram.com', 'instagr.am',
    'twitter.com', 'x.com',
    'linkedin.com',
    'youtube.com', 'youtu.be',
    'tiktok.com',
    'pinterest.com', 'pin.it',
    'reddit.com',
    'threads.net',
    'tumblr.com',
    'flickr.com',
    'snapchat.com',
    't.me', 'telegram.org', 'telegram.me',
    'wa.me', 'whatsapp.com',
    'line.me',
    'viber.com',
    'zalo.me',
    'wechat.com',
    'google.com', 'goo.gl',
    'naver.com', 'daum.net', 'kakao.com',
    'yelp.com', 'tripadvisor.com', 'wikipedia.org',
    'wix.com', 'wixsite.com', 'squarespace.com', 'wordpress.com'
}

def guess_email_from_website(website_url):
    if not website_url or not isinstance(website_url, str):
        return ""
    url = website_url.strip()
    if not url:
        return ""
    if not url.startswith(('http://', 'https://')):
        url = 'http://' + url
    try:
        parsed = urllib.parse.urlparse(url)
        netloc = parsed.netloc.strip().lower()
        if not netloc:
            return ""
        if ':' in netloc:
            netloc = netloc.split(':')[0]
        netloc = re.sub(r'^www\d*\.', '', netloc)
        if not netloc or '.' not in netloc:
            return ""
        for social in SOCIAL_DOMAINS:
            if netloc == social or netloc.endswith('.' + social):
                return ""
        if re.match(r'^[a-z0-9][a-z0-9\.\-]*\.[a-z]{2,}$', netloc):
            if netloc.endswith('.') or re.match(r'^\d+\.\d+\.\d+\.\d+$', netloc):
                return ""
            return f"info@{netloc}"
        return ""
    except Exception:
        return ""

def clean_emails(email_str):
    if not email_str:
        return ""
    parts = [x.strip() for x in email_str.split(",") if x.strip()]
    cleaned = []
    for email in parts:
        domain = email.split("@")[-1].lower()
        if not any(junk in domain for junk in JUNK_DOMAINS) and not domain.endswith(".js"):
            cleaned.append(email)
    return ", ".join(cleaned)

def main():
    input_file = "kofa_members_with_emails.csv"
    output_file = "kofa_members_with_emails_cleaned.csv"
    output_xlsx = "kofa_members_with_emails_cleaned.xlsx"
    
    if not os.path.exists(input_file):
        print(f"[-] File {input_file} not found!")
        return
        
    with open(input_file, mode="r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames
        
    cleaned_count = 0
    valid_emails_count = 0
    
    for row in rows:
        original = row.get("Email", "")
        cleaned = clean_emails(original)
        
        # If we removed some junk emails, track it
        if original and not cleaned:
            cleaned_count += 1
            row["Website"] = ""  # Clear the website too since it was a junk website
        
        if not cleaned:
            website = row.get("Website", "")
            guessed = guess_email_from_website(website)
            if guessed:
                cleaned = guessed

        row["Email"] = cleaned
        if cleaned:
            valid_emails_count += 1
            
    print(f"[*] Cleaned {cleaned_count} rows with third-party emails.")
    print(f"[*] Remaining valid company emails: {valid_emails_count}")
    
    # Save CSV
    with open(output_file, mode="w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"[+] Saved cleaned CSV: {os.path.abspath(output_file)}")
    
    # Save Excel
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cleaned KOFA Emails"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws.append(fieldnames)
        for col_num, header in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        for r_idx, row in enumerate(rows, 1):
            row_data = [row.get(fn, "") for fn in fieldnames]
            ws.append(row_data)
            row_num = r_idx + 1
            
            is_even = (r_idx % 2 == 0)
            for col_num in range(1, len(fieldnames) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_all
                if is_even:
                    cell.fill = fill_zebra
                
                if col_num in [1, 4, 6]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws.row_dimensions[1].height = 28
        for r in range(2, len(rows) + 2):
            ws.row_dimensions[r].height = 20
            
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(output_xlsx)
        print(f"[+] Saved cleaned Excel: {os.path.abspath(output_xlsx)}")
    except Exception as e:
        print(f"[-] Error writing Excel: {e}")

if __name__ == "__main__":
    main()
