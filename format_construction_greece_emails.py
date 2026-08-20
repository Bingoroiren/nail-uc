"""
format_construction_greece_emails.py
=====================================
Formats the "Raw" sheet in "Xây dựng Hy lạp.xlsx" into the standard cold-mail template.

Output columns:
  No. | Công ty | Chức danh | Người liên hệ | SĐT | Liên Hệ | Email |
  Liên Hệ mail | Địa chỉ | Lương | Ngày đăng | Hạn tuyển |
  Check gửi | Last Subject | Last Body HTML | Trạng thái Reply |
  Lần Follow-up | Ngày Follow-up gần nhất | Mailbox đã dùng | Category
"""
import os
import re
import csv
import shutil
import sys
import openpyxl

# Configure stdout/stderr UTF-8 encoding on Windows to prevent UnicodeEncodeError in console print statements
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

INPUT_XLSX = r"d:\glc\nail uc\Xây dựng Hy lạp.xlsx"
if len(sys.argv) > 1:
    INPUT_XLSX = sys.argv[1]

BACKUP_XLSX = INPUT_XLSX.replace(".xlsx", "_backup.xlsx")
FORMATTED_CSV = INPUT_XLSX.replace(".xlsx", "_formatted.csv")
FORMATTED_XLSX = INPUT_XLSX.replace(".xlsx", "_formatted.xlsx")
FORMATTED_V2_XLSX = INPUT_XLSX.replace(".xlsx", "_formatted_v2.xlsx")

GENERIC_DOMAINS = {
    'gmail.com', 'yahoo.com', 'hotmail.com', 'outlook.com', 'live.com', 
    'aol.com', 'icloud.com', 'mail.com', 'ymail.com', 'msn.com', 
    'wix.com', 'squarespace.com', 'wordpress.com', 'wixpress.com'
}

BUSINESS_PREFIXES = {
    'info', 'hello', 'contact', 'office', 'admin', 'sales', 
    'enquiries', 'manager', 'service', 'gemi', 'build', 'builder', 'construction'
}

BAD_KEYWORDS = {
    'wix', 'support', 'no-reply', 'noreply', 'test', 'example', 'domain'
}

def clean_and_score_email(email_str):
    if not email_str:
        return ""
    
    emails = []
    for part in re.split(r'[,\s;]+', str(email_str)):
        clean_part = part.strip().lower()
        if '@' in clean_part:
            emails.append(clean_part)
            
    if not emails:
        return ""
        
    best_email = emails[0]
    best_score = -9999
    
    for email in emails:
        score = 0
        try:
            local_part, domain = email.split('@', 1)
        except ValueError:
            continue
            
        if domain not in GENERIC_DOMAINS:
            score += 10
            
        if any(prefix in local_part for prefix in BUSINESS_PREFIXES):
            score += 5
            
        if any(bad in local_part for bad in BAD_KEYWORDS) or any(bad in domain for bad in BAD_KEYWORDS):
            score -= 20
            
        score -= len(email) * 0.01
        
        if score > best_score:
            best_score = score
            best_email = email
            
    return best_email

def normalize_name(name):
    if not name:
        return ""
    name_clean = str(name).lower().strip()
    name_clean = re.sub(r'[^a-z0-9\sα-ωά-ώ]', '', name_clean) # Keep Latin and Greek characters
    name_clean = " ".join(name_clean.split())
    return name_clean

def normalize_phone(phone_str):
    if not phone_str:
        return ""
    if isinstance(phone_str, (int, float)):
        phone_str = str(int(phone_str))
    else:
        phone_str = str(phone_str).strip()
    
    phone_clean = re.sub(r'[^0-9]', '', phone_str)
    if not phone_clean:
        return ""
        
    # Greece phone numbers: local is 10 digits
    if phone_clean.startswith('0030') and len(phone_clean) > 10:
        phone_clean = phone_clean[4:]
    elif phone_clean.startswith('30') and len(phone_clean) > 10:
        phone_clean = phone_clean[2:]
        
    return phone_clean

def main():
    if not os.path.exists(INPUT_XLSX):
        print(f"Error: Input file {INPUT_XLSX} does not exist.")
        return
        
    print(f"[*] Backing up original XLSX to {BACKUP_XLSX}...")
    shutil.copy2(INPUT_XLSX, BACKUP_XLSX)
    
    print(f"[*] Loading original XLSX: {INPUT_XLSX}...")
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    if 'Raw' not in wb.sheetnames:
        print("Error: Sheet 'Raw' not found in Excel file.")
        return
    sheet = wb['Raw']
    
    # Read rows
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter)
    
    print(f"[+] Headers found: {headers}")
    
    # Map column headers to index
    # Expected columns: ['Tên công ty', 'Tên tiếng Anh', None, 'Điện thoại', 'Email', 'Link chi tiết']
    idx_map = {
        'company_el': 0, # Tên công ty
        'company_en': 1, # Tên tiếng Anh
        'phone': 3,      # Điện thoại
        'email': 4,      # Email
        'link': 5        # Link chi tiết
    }
    
    rows = []
    for r in rows_iter:
        if not r or all(cell is None for cell in r):
            continue
        
        comp_el = r[idx_map['company_el']] if len(r) > idx_map['company_el'] and r[idx_map['company_el']] is not None else ""
        comp_en = r[idx_map['company_en']] if len(r) > idx_map['company_en'] and r[idx_map['company_en']] is not None else ""
        phone = r[idx_map['phone']] if len(r) > idx_map['phone'] and r[idx_map['phone']] is not None else ""
        email = r[idx_map['email']] if len(r) > idx_map['email'] and r[idx_map['email']] is not None else ""
        link = r[idx_map['link']] if len(r) > idx_map['link'] and r[idx_map['link']] is not None else ""
        
        rows.append({
            'Tên công ty': comp_el,
            'Tên tiếng Anh': comp_en,
            'Điện thoại': phone,
            'Email': email,
            'Link chi tiết': link
        })
        
    print(f"[+] Loaded {len(rows)} rows.")
    
    # Step 1: Clean and score emails
    for r in rows:
        r['Best_Email'] = clean_and_score_email(r['Email'])
        
    # Step 2: Deduplicate by company name (English first, fallback to Greek)
    grouped_rows = {}
    for r in rows:
        comp_name = r['Tên tiếng Anh'] if r['Tên tiếng Anh'] else r['Tên công ty']
        norm_name = normalize_name(comp_name)
        if not norm_name:
            continue
        grouped_rows.setdefault(norm_name, []).append(r)
        
    deduplicated_rows = []
    for norm_name, r_list in grouped_rows.items():
        def sort_rep(item):
            has_email = 1 if item['Best_Email'] else 0
            has_link = 1 if str(item['Link chi tiết']).strip() else 0
            return (-has_email, -has_link)
        r_list.sort(key=sort_rep)
        deduplicated_rows.append(r_list[0])
        
    print(f"[+] Deduplicated from {len(rows)} to {len(deduplicated_rows)} unique company names.")
    
    # Step 3: Separate and deduplicate by email and phone number
    seen_emails = set()
    seen_phones = set()
    
    with_email = []
    without_email = []
    
    for r in deduplicated_rows:
        email = r['Best_Email']
        if not email:
            continue
        clean_email = email.strip().lower()
        phone = r['Điện thoại']
        clean_phone = normalize_phone(phone)
        
        if clean_email in seen_emails:
            continue
        if clean_phone and clean_phone in seen_phones:
            continue
            
        seen_emails.add(clean_email)
        if clean_phone:
            seen_phones.add(clean_phone)
        with_email.append(r)
        
    for r in deduplicated_rows:
        email = r['Best_Email']
        if email:
            continue
        phone = r['Điện thoại']
        clean_phone = normalize_phone(phone)
        
        if clean_phone and clean_phone in seen_phones:
            continue
        if clean_phone:
            seen_phones.add(clean_phone)
        without_email.append(r)
        
    print(f"[+] Found {len(with_email)} rows with email, {len(without_email)} rows without email (after removing duplicate emails and phone numbers).")
    final_sorted_rows = with_email + without_email
    
    # Step 4: Reformat to standard cold mail template columns
    fieldnames = [
        "No.", "Công ty", "Chức danh", "Người liên hệ", "SĐT", "Liên Hệ", 
        "Email", "Liên Hệ mail", "Địa chỉ", "Lương", "Ngày đăng", "Hạn tuyển", 
        "Check gửi", "Last Subject", "Last Body HTML", "Trạng thái Reply", 
        "Lần Follow-up", "Ngày Follow-up gần nhất", "Mailbox đã dùng", "Category"
    ]
    
    output_rows = []
    for i, r in enumerate(final_sorted_rows, 1):
        comp_name = r['Tên tiếng Anh'] if r['Tên tiếng Anh'] else r['Tên công ty']
        if not comp_name:
            comp_name = ""
        comp_name = str(comp_name).strip()
        
        phone = r['Điện thoại']
        if phone:
            raw_digits = normalize_phone(phone)
            phone_val = f"'{raw_digits}"
        else:
            phone_val = ""
            
        out_row = {
            "No.": i,
            "Công ty": comp_name,
            "Chức danh": "",
            "Người liên hệ": "",
            "SĐT": phone_val,
            "Liên Hệ": str(r['Link chi tiết'] or "").strip(),
            "Email": r['Best_Email'],
            "Liên Hệ mail": "",
            "Địa chỉ": "",
            "Lương": "",
            "Ngày đăng": "",
            "Hạn tuyển": "",
            "Check gửi": "",
            "Last Subject": "",
            "Last Body HTML": "",
            "Trạng thái Reply": "",
            "Lần Follow-up": 0,
            "Ngày Follow-up gần nhất": "",
            "Mailbox đã dùng": "",
            "Category": "Xây dựng Hy Lạp"
        }
        output_rows.append(out_row)
        
    # Write CSV
    print(f"[*] Writing to CSV format: {FORMATTED_CSV}...")
    with open(FORMATTED_CSV, mode='w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(output_rows)
        
    # Write XLSX
    target_xlsx = FORMATTED_XLSX
    print(f"[*] Writing to XLSX format: {target_xlsx}...")
    try:
        out_wb = openpyxl.Workbook()
        out_sheet = out_wb.active
        out_sheet.title = "Raw"
        
        # Write headers
        out_sheet.append(fieldnames)
        
        # Write data rows
        for o_row in output_rows:
            row_data = [o_row[k] for k in fieldnames]
            out_sheet.append(row_data)
            
        out_wb.save(target_xlsx)
    except PermissionError:
        target_xlsx = FORMATTED_V2_XLSX
        print(f"[!] {FORMATTED_XLSX} is locked by Excel. Writing to {FORMATTED_V2_XLSX} instead...")
        out_wb = openpyxl.Workbook()
        out_sheet = out_wb.active
        out_sheet.title = "Raw"
        
        out_sheet.append(fieldnames)
        for o_row in output_rows:
            row_data = [o_row[k] for k in fieldnames]
            out_sheet.append(row_data)
        out_wb.save(target_xlsx)
        
    # Try to copy/overwrite the original INPUT_XLSX
    print(f"[*] Attempting to update original {INPUT_XLSX}...")
    try:
        shutil.copy2(target_xlsx, INPUT_XLSX)
        print("[SUCCESS] Successfully updated original file!")
    except PermissionError:
        print("\n[WARNING] Permission denied! The original file is locked by Excel.")
        print(f"[i] The formatted copy is available at: {target_xlsx}")
        print(f"[i] CSV version is available at: {FORMATTED_CSV}")

if __name__ == "__main__":
    main()
