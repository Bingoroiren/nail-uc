# scraper_naver.py
# Scrapes recruitment and staffing agencies from Naver Map (Mobile version) using Playwright

import asyncio
import csv
import os
import random
import sys
import json
import urllib.parse
from playwright.async_api import async_playwright
from playwright_stealth import stealth_async
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Import local Korean config and locations
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import config_kr
import locations_kr

# Set console output encoding to UTF-8
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

PROGRESS_FILE = os.path.join(os.path.dirname(config_kr.OUTPUT_CSV), "scraping_progress_naver.json")

def safe_print(msg):
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        try:
            print(msg.encode(sys.stdout.encoding or 'utf-8', errors='replace').decode(sys.stdout.encoding or 'utf-8'), flush=True)
        except Exception:
            print(msg.encode('ascii', errors='replace').decode('ascii'), flush=True)

def load_completed_scans():
    completed = set()
    if os.path.exists(PROGRESS_FILE):
        try:
            with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return set((item[0].lower(), item[1].lower(), item[2].lower()) for item in data.get("completed", []))
        except Exception as e:
            safe_print(f"[-] Error loading progress file: {e}")
            
    # Fallback: Initialize from existing CSV if progress.json is missing
    if os.path.exists(config_kr.OUTPUT_CSV):
        try:
            safe_print("[*] Progress file not found. Initializing from existing CSV data...")
            completed_list = []
            with open(config_kr.OUTPUT_CSV, mode='r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    loc = row.get("Khu vực tìm kiếm / Search Location", "")
                    kw = row.get("Từ khóa tìm kiếm / Search Keyword", "")
                    if loc and kw:
                        parts = loc.split(" ", 1)
                        city = parts[0].strip().lower()
                        district = parts[1].strip().lower() if len(parts) > 1 else ""
                        pair = (city, district, kw.strip().lower())
                        if pair not in completed:
                            completed.add(pair)
                            completed_list.append([city, district, kw.strip()])
            with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
                json.dump({"completed": completed_list}, f, indent=4, ensure_ascii=False)
            safe_print(f"[+] Loaded {len(completed)} completed scans from CSV.")
        except Exception as e:
            safe_print(f"[-] Error initializing progress from CSV: {e}")
    return completed

def save_completed_scan(city, district, keyword):
    completed_list = []
    if os.path.exists(PROGRESS_FILE):
        try:
            with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                completed_list = data.get("completed", [])
        except:
            pass
    pair = [city, district, keyword]
    if pair not in completed_list:
        completed_list.append(pair)
        try:
            with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
                json.dump({"completed": completed_list}, f, indent=4, ensure_ascii=False)
        except Exception as e:
            safe_print(f"[-] Error saving progress file: {e}")

def get_next_stt():
    if not os.path.exists(config_kr.OUTPUT_CSV):
        return 1
    try:
        with open(config_kr.OUTPUT_CSV, mode='r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = list(reader)
            return len(rows) # Headers + rows, so len(rows) is next index
    except:
        return 1

def append_to_csv(row_dict):
    file_exists = os.path.isfile(config_kr.OUTPUT_CSV)
    try:
        with open(config_kr.OUTPUT_CSV, mode='a', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=row_dict.keys())
            if not file_exists:
                writer.writeheader()
            writer.writerow(row_dict)
    except Exception as e:
        safe_print(f"[-] Failed to write row to CSV: {e}")

async def get_official_website(context, place_url):
    """Navigates to the Naver Place page to extract the business's actual official website URL."""
    if not place_url:
        return ""
    
    page = None
    try:
        page = await context.new_page()
        await stealth_async(page)
        await page.goto(place_url, timeout=12000, wait_until="commit")
        await page.wait_for_timeout(1000)
        
        # Scrape all links and find external domain link
        links = await page.locator("a").all()
        for link in links:
            href = await link.get_attribute("href")
            if href and href.startswith(("http://", "https://")):
                href_lower = href.lower()
                # Filter out sharing portals and naver internal domains
                if not any(x in href_lower for x in ["naver.com", "facebook.com/sharer", "twitter.com", "instagram.com"]):
                    return href
    except Exception:
        pass
    finally:
        if page:
            await page.close()
    return ""

async def main_async():
    completed_scans = load_completed_scans()
    
    headers = [
        "STT / No.",
        "Tên công ty / Company Name",
        "Danh mục / Category",
        "Số điện thoại / Phone",
        "Địa chỉ / Address",
        "Website",
        "Naver Place Link",
        "Khu vực tìm kiếm / Search Location",
        "Từ khóa tìm kiếm / Search Keyword"
    ]
    
    # Initialize empty CSV with headers if it doesn't exist
    if not os.path.exists(config_kr.OUTPUT_CSV):
        with open(config_kr.OUTPUT_CSV, mode="w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
    total_locations = len(locations_kr.LOCATIONS)
    total_keywords = len(config_kr.KEYWORDS)
    
    safe_print(f"[*] Starting Naver Map Mobile Scraper...")
    safe_print(f"[*] Total Locations: {total_locations}, Keywords: {total_keywords}")
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        # Mobile Device emulation
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (iPhone; CPU iPhone OS 14_8 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.2 Mobile/15E148 Safari/604.1",
            viewport={"width": 375, "height": 812},
            locale="ko-KR",
            extra_http_headers={"Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7"}
        )
        page = await context.new_page()
        await stealth_async(page)
        
        loc_idx = 0
        for loc in locations_kr.LOCATIONS:
            loc_idx += 1
            city = loc["city"]
            district = loc["district"]
            
            for kw in config_kr.KEYWORDS:
                scan_key = (city.lower(), district.lower(), kw.lower())
                if scan_key in completed_scans:
                    continue
                    
                query = f"{city} {district} {kw}"
                safe_print(f"\n[Location {loc_idx}/{total_locations}] Searching for: '{query}'...")
                
                search_url = f"https://m.map.naver.com/search2/search.naver?query={urllib.parse.quote(query)}"
                try:
                    await page.goto(search_url, timeout=30000, wait_until="commit")
                    await page.wait_for_timeout(3000)
                    
                    # Check "No results found" text
                    body_text = await page.locator("body").inner_text()
                    if "검색결과가 없습니다" in body_text or "검색 결과가 없습니다" in body_text:
                        safe_print("  [-] No search results found on Naver Map.")
                        save_completed_scan(city, district, kw)
                        continue
                        
                    # Scroll to bottom to lazy load results (mobile Naver Map lists up to 75 results on scroll)
                    safe_print("  [*] Scrolling to load all places...")
                    for scroll_step in range(6):
                        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                        await page.wait_for_timeout(800)
                        
                    # Find all result list items (elements with class containing list_item)
                    items = await page.locator("ul.search_list > li, li[class*='_list_item_']").all()
                    safe_print(f"  [+] Found {len(items)} items in result list.")
                    
                    for i_idx, item in enumerate(items, 1):
                        # 1. Extract Title
                        title_el = item.locator("strong").first
                        if await title_el.count() == 0:
                            continue
                        name = await title_el.inner_text()
                        name = name.strip() if name else ""
                        if not name:
                            continue
                            
                        # 2. Extract Category
                        category = ""
                        cat_el = item.locator("em").first
                        if await cat_el.count() > 0:
                            category = await cat_el.inner_text()
                            category = category.strip() if category else ""
                            
                        # Filter by target category tags
                        if category and any(tag in category for tag in config_kr.TARGET_TAGS):
                            # 3. Extract Address
                            address = ""
                            addr_el = item.locator("button[class*='_item_address_']").first
                            if await addr_el.count() > 0:
                                address = await addr_el.inner_text()
                                address = address.replace("주소보기", "").strip()
                                
                            # 4. Extract Phone
                            phone = ""
                            tel_link = item.locator("a[href^='tel:']").first
                            if await tel_link.count() > 0:
                                phone = await tel_link.get_attribute("href")
                                phone = phone.replace("tel:", "").strip()
                                
                            # 5. Extract Naver Place Link
                            place_url = ""
                            place_link_el = item.locator("a[href*='place.naver.com']").first
                            if await place_link_el.count() > 0:
                                place_url = await place_link_el.get_attribute("href")
                                
                            # 6. Crawl Naver Place page to get official website
                            website = ""
                            if place_url:
                                website = await get_official_website(context, place_url)
                                
                            # Save matching recruiter
                            row_data = {
                                "STT / No.": get_next_stt(),
                                "Tên công ty / Company Name": name,
                                "Danh mục / Category": category,
                                "Số điện thoại / Phone": f"'{phone}" if phone else "",
                                "Địa chỉ / Address": address,
                                "Website": website,
                                "Naver Place Link": place_url,
                                "Khu vực tìm kiếm / Search Location": f"{city} {district}",
                                "Từ khóa tìm kiếm / Search Keyword": kw
                            }
                            append_to_csv(row_data)
                            safe_print(f"    [+] Saved matching agency ({i_idx}/{len(items)}): {name} ({category}) -> Phone: {phone}")
                            
                            # Random delay to prevent rate limit
                            await page.wait_for_timeout(random.uniform(500, 1500))
                            
                except Exception as e:
                    safe_print(f"  [-] Error scanning Naver Map query: {e}")
                    
                # Scan completed successfully
                save_completed_scan(city, district, kw)
                completed_scans.add(scan_key)
                await page.wait_for_timeout(random.uniform(2000, 4000))
                
        await browser.close()
        
    # Write to final Excel
    safe_print(f"\n[*] Writing results to Excel: {config_kr.OUTPUT_XLSX}...")
    try:
        # Load CSV
        with open(config_kr.OUTPUT_CSV, mode="r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            fieldnames = reader.fieldnames
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Naver Map Recruiter"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws.append(fieldnames)
        for col_num, header_name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        for r_idx, row in enumerate(rows, 1):
            row_data = [row.get(fn, "") for fn in fieldnames]
            ws.append(row_data)
            row_num = r_idx + 1
            
            is_even = (r_idx % 2 == 0)
            for col_num in range(1, len(fieldnames) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_all
                if is_even:
                    cell.fill = fill_zebra
                
                if col_num in [1, 4, 8, 9]: # STT, Phone, Location, Keyword
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws.row_dimensions[1].height = 28
        for r in range(2, len(rows) + 2):
            ws.row_dimensions[r].height = 20
            
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(config_kr.OUTPUT_XLSX)
        safe_print(f"[+] Saved final Excel: {os.path.abspath(config_kr.OUTPUT_XLSX)}")
    except Exception as e:
        safe_print(f"[-] Error writing Excel: {e}")

if __name__ == "__main__":
    asyncio.run(main_async())
