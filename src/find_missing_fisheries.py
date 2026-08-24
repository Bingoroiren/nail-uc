import asyncio
import csv
import json
import os
import random
import re
import sys
import urllib.parse
import openpyxl
from playwright.async_api import async_playwright

# Import local configuration
import config_fisheries_lv

# Force UTF-8 output
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

xlsx_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "8.sekcija - Zvejas produktu apstrādes uzņēmumi.xlsx")
csv_path = config_fisheries_lv.OUTPUT_CSV

def clean_display_name(name):
    if not name:
        return ""
    name = name.strip()
    # Normalize quotes
    name = name.replace('„', '"').replace('“', '"').replace('”', '"').replace('„', '"')
    return name

def clean_search_query(name):
    # Remove quotes and clean for search query
    name = name.replace('"', '').replace('„', '').replace('“', '').replace('”', '')
    # Remove extra spaces
    name = re.sub(r'\s+', ' ', name).strip()
    return f"{name}, Latvia"

def extract_place_id(url):
    if not url:
        return ""
    match = re.search(r'1s(0x[0-9a-fA-F]+:0x[0-9a-fA-F]+)', url)
    if match:
        return match.group(1).lower()
    return url.split('?')[0].lower()

def is_latvia_address(address):
    if not address:
        return False
    addr_lower = address.lower()
    if "latvia" in addr_lower or "latvija" in addr_lower or " lv-" in addr_lower or addr_lower.startswith("lv-") or ", lv " in addr_lower:
        return True
    lv_cities = ["rīga", "riga", "liepāja", "liepaja", "daugavpils", "jelgava", "jūrmala", "jurmala", "ventspils", "rēzekne", "rezekne", "valmiera", "jēkabpils", "jekabpils", "ogre", "salaspils", "salacgrīva", "roja", "engure", "pāvilosta", "kolka"]
    if any(city in addr_lower for city in lv_cities):
        return True
    if re.search(r'\blv-\d{4}\b', addr_lower) or re.search(r'\b\d{4}\b', addr_lower):
        return True
    return False

def extract_coords_from_url(url):
    if not url:
        return None
    match = re.search(r'!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)', url)
    if match:
        return float(match.group(1)), float(match.group(2))
    match = re.search(r'@(-?\d+\.\d+),(-?\d+\.\d+)', url)
    if match:
        return float(match.group(1)), float(match.group(2))
    return None

def parse_latvian_address(address):
    if not address:
        return None, None, None
    addr = re.sub(r',\s*Latvia\s*$', '', address, flags=re.IGNORECASE).strip()
    addr = re.sub(r',\s*Latvija\s*$', '', addr, flags=re.IGNORECASE).strip()
    
    postcode_match = re.search(r'\bLV-\d{4}\b', addr, flags=re.IGNORECASE)
    if postcode_match:
        postcode = postcode_match.group(0).upper()
        remaining = addr[:postcode_match.start()].strip().rstrip(',')
        parts = [p.strip() for p in remaining.split(',')]
        suburb = parts[-1] if parts else ""
        return suburb, "LV", postcode
        
    postcode_match = re.search(r'\b(\d{4})$', addr)
    if postcode_match:
        postcode = f"LV-{postcode_match.group(1)}"
        remaining = addr[:postcode_match.start()].strip().rstrip(',')
        parts = [p.strip() for p in remaining.split(',')]
        suburb = parts[-1] if parts else ""
        return suburb, "LV", postcode
        
    parts = [p.strip() for p in addr.split(',')]
    if len(parts) >= 2:
        return parts[-2], "LV", ""
    elif len(parts) == 1:
        return parts[0], "LV", ""
    return None, None, None

def get_scraped_urls():
    scraped_urls = set()
    if os.path.exists(csv_path):
        try:
            with open(csv_path, mode='r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if 'URL' in row and row['URL']:
                        scraped_urls.add(extract_place_id(row['URL']))
        except Exception as e:
            print(f"[-] Error loading existing CSV records: {e}")
    return scraped_urls

def append_to_csv(row_dict):
    file_exists = os.path.isfile(csv_path)
    try:
        with open(csv_path, mode='a', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=row_dict.keys())
            if not file_exists:
                writer.writeheader()
            writer.writerow(row_dict)
    except Exception as e:
        print(f"[-] Failed to write row to CSV: {e}")

async def handle_captcha(page):
    is_captcha = False
    try:
        title = await page.title()
        if "sorry" in title.lower() or "recaptcha" in title.lower() or "captcha" in title.lower():
            is_captcha = True
        elif await page.locator('iframe[src*="recaptcha"]').count() > 0 or await page.locator('div#recaptcha').count() > 0:
            is_captcha = True
    except Exception:
        pass
    if is_captcha:
        print("\n" + "="*60)
        print("[!] IP BLOCK / CAPTCHA DETECTED! Google is blocking automated access.")
        print("[!] Please solve CAPTCHA in browser or turn on VPN/proxy.")
        print("="*60 + "\n")
        sys.stdout.write('\a')
        sys.stdout.flush()
        # Pause script to let user solve it or switch IP
        input("Press Enter after solving CAPTCHA / changing VPN to resume...")

async def bypass_consent_screen(page):
    try:
        consent_buttons = page.locator('button:has-text("Accept all"), button:has-text("Agree"), button:has-text("I agree"), button:has-text("Accept"), button:has-text("Accept details"), button:has-text("Piekrist visam"), button:has-text("Es piekrītu")')
        if await consent_buttons.count() > 0:
            print("[*] Bypassing Google Consent Screen...")
            await consent_buttons.first.click()
            await page.wait_for_load_state("networkidle")
            await page.wait_for_timeout(1000)
    except Exception:
        pass

async def extract_details(page, url, search_query):
    sel = config_fisheries_lv.SELECTORS
    
    name = ""
    name_loc = page.locator(sel["business_name"])
    if await name_loc.count() > 0:
        name = await name_loc.first.inner_text()
        name = name.strip()
        
    if not name:
        return None
        
    website = ""
    web_loc = page.locator(sel["website"])
    if await web_loc.count() > 0:
        website = await web_loc.first.get_attribute("href")
        if website:
            website = website.strip()
            
    phone = ""
    phone_loc = page.locator(sel["phone"])
    if await phone_loc.count() > 0:
        phone_attr = await phone_loc.first.get_attribute("data-item-id")
        if phone_attr:
            phone = phone_attr.replace("phone:tel:", "").strip()
            
    address = ""
    addr_loc = page.locator(sel["address"])
    if await addr_loc.count() > 0:
        addr_label = await addr_loc.first.get_attribute("aria-label")
        if addr_label:
            address = addr_label.replace("Address:", "").replace("Adrese:", "").strip()
        else:
            address = await addr_loc.first.inner_text()
            address = address.strip()
            
    if not is_latvia_address(address):
        print(f"    [-] Skipping: Address '{address}' is not in Latvia.")
        return None
        
    actual_lat, actual_lng = 56.95, 24.10  # Default Riga
    coords = extract_coords_from_url(url)
    if coords:
        actual_lat, actual_lng = coords
        
    actual_suburb, actual_state, postcode = parse_latvian_address(address)
    if not actual_state:
        actual_suburb = "Latvia"
        actual_state = "LV"
        
    rating = ""
    reviews_count = ""
    rating_loc = page.locator('div.F7nice')
    if await rating_loc.count() > 0:
        span_rating = rating_loc.first.locator('span[aria-hidden="true"]')
        if await span_rating.count() > 0:
            rating = await span_rating.first.inner_text()
            rating = rating.strip()
            
        span_reviews = rating_loc.first.locator('span[aria-label*="atsauksme"], span[aria-label*="review"]')
        if await span_reviews.count() > 0:
            reviews_text = await span_reviews.first.get_attribute("aria-label")
            if reviews_text:
                match = re.search(r'\d+', reviews_text.replace(" ", "").replace(",", ""))
                if match:
                    reviews_count = match.group()
            else:
                reviews_text = await span_reviews.first.inner_text()
                match = re.search(r'\d+', reviews_text.replace(" ", "").replace(",", ""))
                if match:
                    reviews_count = match.group()
                    
    permanently_closed = "No"
    try:
        closed_loc = page.locator('span:has-text("Permanently closed"), span:has-text("Slēgts uz visiem laikiem"), span:has-text("Đóng cửa vĩnh viễn")')
        if await closed_loc.count() > 0:
            permanently_closed = "Yes"
    except Exception:
        pass

    category = ""
    try:
        category_loc = page.locator(sel["category"])
        if await category_loc.count() > 0:
            category = await category_loc.first.inner_text()
            category = category.strip()
    except Exception:
        pass

    record = {
        "Name": name,
        "Website": website,
        "Phone": phone,
        "Address": address,
        "Rating": rating,
        "Reviews_Count": reviews_count,
        "State": actual_state,
        "Location_Name": actual_suburb,
        "Latitude": actual_lat,
        "Longitude": actual_lng,
        "Search_Query": search_query,
        "URL": url,
        "Permanently_Closed": permanently_closed,
        "Category": category if category else "Zivju apstrādes uzņēmums",
    }
    return record

# Helper to normalize name for comparison
def clean_name_for_comparison(name):
    if not name:
        return ""
    name = name.lower()
    name = name.replace('sabiedrība ar ierobežotu atbildību', '')
    name = name.replace('komandītsabiedrība', '')
    name = name.replace('akciju sabiedrība', '')
    name = name.replace('sia', '')
    name = name.replace('as', '')
    name = name.replace('ks', '')
    name = name.replace('i/u', '')
    name = name.replace('iu', '')
    name = name.replace('"', '').replace('“', '').replace('”', '').replace('„', '')
    name = re.sub(r'[^a-z0-9āčēģīķļņšūž]', '', name)
    return name.strip()

async def main():
    print("=============================================================")
    print("      GOOGLE MAPS SEARCH FOR MISSING FISH PROCESSING FACTORIES ")
    print("=============================================================")
    
    # 1. Read Excel companies
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb['SEKC 8']
    excel_companies = []
    for r_idx in range(8, sheet.max_row + 1):
        reg_num = sheet.cell(row=r_idx, column=2).value
        name = sheet.cell(row=r_idx, column=3).value
        if name and reg_num:
            excel_companies.append({
                'reg_num': reg_num.strip(),
                'name': name.strip()
            })
            
    print(f"[*] Loaded {len(excel_companies)} companies from Excel.")
    
    # 2. Read existing CSV names
    scraped_urls = get_scraped_urls()
    csv_names = set()
    if os.path.exists(csv_path):
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for r in reader:
                name = r.get('Name', '').strip()
                if name:
                    csv_names.add(clean_name_for_comparison(name))
                    
    print(f"[*] Loaded {len(csv_names)} existing names from CSV.")
    
    # 3. Identify missing companies
    missing_companies = []
    for c in excel_companies:
        c_clean = clean_name_for_comparison(c['name'])
        found = False
        if c_clean in csv_names:
            found = True
        else:
            for name in csv_names:
                if name and (name in c_clean or c_clean in name):
                    found = True
                    break
        if not found:
            missing_companies.append(c)
            
    print(f"[*] Found {len(missing_companies)} missing companies to search on Google Maps.")
    
    if not missing_companies:
        print("[SUCCESS] All companies are already in the CSV file!")
        return

    # 4. Launch Playwright
    async with async_playwright() as p:
        browser = None
        for channel in ["chrome", "msedge", None]:
            try:
                chan_str = f"channel '{channel}'" if channel else "default Chromium"
                print(f"[*] Attempting to launch browser with {chan_str}...")
                launch_args = {
                    "headless": False,  # Show browser so user can see it
                    "slow_mo": config_fisheries_lv.SLOW_MO,
                    "args": [
                        "--disable-blink-features=AutomationControlled",
                        "--lang=lv-LV,lv"
                    ]
                }
                if channel:
                    launch_args["channel"] = channel
                browser = await p.chromium.launch(**launch_args)
                print(f"[+] Successfully launched browser using {chan_str}!")
                break
            except Exception as e:
                print(f"[-] Failed to launch with channel '{channel}': {e}")
                
        if not browser:
            print("[!] Could not launch any browser. Exiting.")
            return
            
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            locale="lv-LV",
            viewport={"width": 1280, "height": 800},
            geolocation={"latitude": 56.9475, "longitude": 24.1069},
            permissions=["geolocation"]
        )
        page = await context.new_page()
        
        # Load Google Maps homepage once to solve consent banner
        await page.goto("https://www.google.com/maps?hl=lv", timeout=30000)
        await page.wait_for_timeout(2000)
        await bypass_consent_screen(page)
        
        success_count = 0
        failed_count = 0
        
        for idx, comp in enumerate(missing_companies, 1):
            name = comp['name']
            query = clean_search_query(name)
            print(f"\n[{idx}/{len(missing_companies)}] Searching: '{query}'")
            
            # Go to search URL
            query_encoded = urllib.parse.quote_plus(query)
            search_url = f"https://www.google.com/maps/search/{query_encoded}?hl=lv"
            
            try:
                await page.goto(search_url, timeout=30000, wait_until="domcontentloaded")
                await page.wait_for_timeout(2000)
                await handle_captcha(page)
                await bypass_consent_screen(page)
                
                current_url = page.url
                place_saved = False
                
                # Check if it loaded the place details panel directly
                if "/maps/place/" in current_url:
                    place_id = extract_place_id(current_url)
                    if place_id not in scraped_urls:
                        record = await extract_details(page, current_url, query)
                        if record:
                            append_to_csv(record)
                            scraped_urls.add(place_id)
                            place_saved = True
                            print(f"    -> [SUCCESS] Direct Match: {record['Name']} | Phone: {record['Phone']} | Web: {record['Website']}")
                    else:
                        print("    [-] Already scraped (URL matched).")
                        place_saved = True
                        
                else:
                    # It's a list or no results page. Let's see if any matching element is listed
                    link_selector = config_fisheries_lv.SELECTORS["listing_link"]
                    listings_count = await page.locator(link_selector).count()
                    
                    if listings_count > 0:
                        # Click the first result
                        first_item = page.locator(link_selector).first
                        expected_name = await first_item.get_attribute("aria-label")
                        await first_item.scroll_into_view_if_needed()
                        await first_item.click()
                        
                        # Wait for name header to appear in detail panel
                        name_matched = False
                        for _ in range(15):
                            h1_locator = page.locator(config_fisheries_lv.SELECTORS["business_name"])
                            if await h1_locator.count() > 0:
                                name_matched = True
                                break
                            await page.wait_for_timeout(200)
                            
                        if name_matched:
                            current_url = page.url
                            place_id = extract_place_id(current_url)
                            if place_id not in scraped_urls:
                                record = await extract_details(page, current_url, query)
                                if record:
                                    append_to_csv(record)
                                    scraped_urls.add(place_id)
                                    place_saved = True
                                    print(f"    -> [SUCCESS] List Match: {record['Name']} | Phone: {record['Phone']} | Web: {record['Website']}")
                            else:
                                print("    [-] Already scraped (URL matched).")
                                place_saved = True
                
                if place_saved:
                    success_count += 1
                else:
                    print("    [-] Not found / no exact matches on Google Maps.")
                    failed_count += 1
                    
            except Exception as e:
                print(f"    [!] Error during search: {e}")
                failed_count += 1
                
            # Random delay between queries to look human
            await page.wait_for_timeout(random.uniform(1500, 3000))
            
        print(f"\n=============================================================")
        print(f"  FINISHED: Found {success_count} new places. {failed_count} not found.")
        print(f"  Total records in CSV: {len(get_scraped_urls())}")
        print("=============================================================")
        
        await browser.close()

if __name__ == "__main__":
    asyncio.run(main())
