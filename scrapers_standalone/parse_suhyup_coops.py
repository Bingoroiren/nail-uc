import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
import re
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def parse_html_table(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        html = f.read()
        
    cooperatives = []
    
    # Find all tr blocks
    tr_blocks = re.findall(r'<tr>([\s\S]*?)</tr>', html)
    
    for tr in tr_blocks:
        tds = re.findall(r'<td[^>]*>([\s\S]*?)</td>', tr)
        if len(tds) >= 5:
            # Clean HTML tags and entities
            stt = re.sub(r'<[^>]+>', '', tds[0]).strip()
            name = re.sub(r'<[^>]+>', '', tds[1]).strip()
            region = re.sub(r'<[^>]+>', '', tds[2]).strip()
            address = re.sub(r'<[^>]+>', '', tds[3]).strip()
            phone = re.sub(r'<[^>]+>', '', tds[4]).strip()
            
            # Remarks column might not exist or be empty
            remarks = ""
            if len(tds) >= 6:
                remarks = re.sub(r'<[^>]+>', '', tds[5]).strip()
                
            cooperatives.append({
                "stt": stt,
                "name": name,
                "region": region,
                "address": address,
                "phone": phone,
                "remarks": remarks
            })
            
    return cooperatives

def main():
    source_file = os.path.join(ROOT_DIR, "data", "raw", "suhyup_coops.xls")
    if not os.path.exists(source_file):
        print(f"[-] Source file {source_file} not found!")
        return
        
    coops = parse_html_table(source_file)
    print(f"[+] Successfully parsed {len(coops)} Suhyup cooperatives.")
    
    headers = [
        "STT / No.",
        "Tên hợp tác xã / Cooperative Name",
        "Khu vực / Region",
        "Địa chỉ / Address",
        "Số điện thoại / Phone",
        "Ghi chú / Remarks"
    ]
    
    csv_file = os.path.join(ROOT_DIR, "data", "raw", "suhyup_coops.csv")
    xlsx_file = os.path.join(ROOT_DIR, "data", "raw", "suhyup_coops.xlsx")
    
    # 1. Write CSV
    print(f"[*] Writing to CSV: {csv_file}")
    try:
        with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for item in coops:
                writer.writerow([
                    item["stt"],
                    item["name"],
                    item["region"],
                    item["address"],
                    item["phone"],
                    item["remarks"]
                ])
        print(f"[+] Saved CSV file successfully: {os.path.abspath(csv_file)}")
    except Exception as e:
        print(f"[-] Error writing CSV: {e}")
        
    # 2. Write Excel
    print(f"[*] Writing to Excel: {xlsx_file}")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Suhyup Cooperatives"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        for idx, item in enumerate(coops, 1):
            row_data = [
                item["stt"],
                item["name"],
                item["region"],
                item["address"],
                item["phone"],
                item["remarks"]
            ]
            ws.append(row_data)
            row_num = idx + 1
            
            is_even = (idx % 2 == 0)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_all
                if is_even:
                    cell.fill = fill_zebra
                
                if col_num in [1, 3, 5]: # STT, Region, Phone
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws.row_dimensions[1].height = 28
        for r in range(2, len(coops) + 2):
            ws.row_dimensions[r].height = 20
            
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(xlsx_file)
        print(f"[+] Saved Excel file successfully: {os.path.abspath(xlsx_file)}")
    except Exception as e:
        print(f"[-] Error writing Excel: {e}")

if __name__ == "__main__":
    main()
