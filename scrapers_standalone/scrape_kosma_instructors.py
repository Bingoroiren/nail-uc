import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
import urllib.request
import urllib.parse
import json
import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def scrape_kosma_members():
    print("[*] Accessing KOSMA Instructor/Member company directory API...")
    url = "http://www.kosma.or.kr/selectCompanyInfoList.do"

    # We fetch all 158 records in a single request (using PAGE_ROW=300 to be safe for future growth)
    data = {
        "PAGE_INDEX": "1",
        "PAGE_ROW": "300",
        "searchString": ""
    }
    data_encoded = urllib.parse.urlencode(data).encode("utf-8")

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest"
    }

    req = urllib.request.Request(url, data=data_encoded, headers=headers, method="POST")

    try:
        with urllib.request.urlopen(req) as response:
            html = response.read().decode("utf-8")
            parsed = json.loads(html)
            
            raw_list = parsed.get("list", [])
            total_count = parsed.get("TOTAL", 0)
            print(f"[+] Successfully fetched {len(raw_list)} records (reported total: {total_count})")
            
            return raw_list
    except Exception as e:
        print(f"[-] Error during request: {e}")
        return []

def format_date(timestamp_ms):
    if not timestamp_ms:
        return ""
    try:
        # Convert milliseconds timestamp to readable date string
        dt = datetime.fromtimestamp(timestamp_ms / 1000.0)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return ""

def main():
    members = scrape_kosma_members()
    if not members:
        print("[-] No data retrieved. Exiting.")
        return

    # Sort by ROW_NUM ascending (or reverse chronological as returned)
    # The list is returned in reverse chronological order (newest first). Let's reverse it to have chronological order or keep as is.
    # We will number them 1 to N in the order of occurrence.
    
    headers = [
        "STT / No.",
        "Tên công ty / Company Name",
        "Người đại diện / Representative",
        "Số điện thoại / Phone",
        "Số fax / Fax",
        "Địa chỉ / Address",
        "Ngày đăng ký / Reg Date"
    ]

    csv_file = os.path.join(ROOT_DIR, "data", "raw", "kosma_instructors.csv")
    xlsx_file = os.path.join(ROOT_DIR, "data", "raw", "kosma_instructors.xlsx")

    # 1. Write CSV with UTF-8-BOM to support Korean characters in Excel
    print(f"[*] Writing to CSV file: {csv_file}...")
    try:
        with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for idx, item in enumerate(members, 1):
                writer.writerow([
                    idx,
                    item.get("company_name", "").strip(),
                    item.get("owner_name", "").strip(),
                    item.get("tel", "").strip(),
                    item.get("fax", "").strip(),
                    item.get("address", "").strip(),
                    item.get("write_date", "").strip() or format_date(item.get("reg_date"))
                ])
        print(f"[+] Saved CSV file successfully: {os.path.abspath(csv_file)}")
    except Exception as e:
        print(f"[-] Error writing CSV: {e}")

    # 2. Write Excel with premium styles using openpyxl
    print(f"[*] Writing to Excel file: {xlsx_file}...")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "KOSMA Members"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        font_bold = Font(name="Calibri", size=11, bold=True)
        
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Deep navy
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
        
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Write headers
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
        
        # Write data rows
        for idx, item in enumerate(members, 1):
            row_data = [
                idx,
                item.get("company_name", "").strip(),
                item.get("owner_name", "").strip(),
                item.get("tel", "").strip(),
                item.get("fax", "").strip(),
                item.get("address", "").strip(),
                item.get("write_date", "").strip() or format_date(item.get("reg_date"))
            ]
            ws.append(row_data)
            row_num = idx + 1
            
            # Apply formatting
            is_even = (idx % 2 == 0)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_all
                
                # Zebra striping
                if is_even:
                    cell.fill = fill_zebra
                
                # Alignments
                if col_num in [1, 7]: # No., Date
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in [4, 5]: # Phone, Fax
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        # Set row heights
        ws.row_dimensions[1].height = 28
        for r in range(2, len(members) + 2):
            ws.row_dimensions[r].height = 20
            
        # Autofit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(xlsx_file)
        print(f"[+] Saved Excel file successfully: {os.path.abspath(xlsx_file)}")
    except Exception as e:
        print(f"[-] Error writing Excel: {e}")

if __name__ == "__main__":
    main()
