import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
import asyncio
import csv
import os
import re
import urllib.parse
from playwright.async_api import async_playwright
from playwright_stealth import stealth_async
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# List of domains to exclude when looking for official websites
EXCLUDE_DOMAINS = [
    "namu.wiki", "wikipedia.org", "naver.com", "daum.net", "tistory.com", 
    "egloos.com", "youtube.com", "jobkorea.co.kr", "saramin.co.kr", "catch.co.kr",
    "people.incruit.com", "nicebizinfo.com", "dart.fss.or.kr", "facebook.com/login",
    "map.naver.com", "place.map.kakao.com", "maps.google.com", "fss.or.kr", "hometax.go.kr",
    "blo.lg", "blog.naver.com", "cafe.naver.com"
]

# Domains that belong to news outlets, directories or software agencies to filter from email results
JUNK_DOMAINS = [
    "news2day.co.kr", "newsprime.co.kr", "insight.co.kr", "siminilbo.co.kr", 
    "naeil.com", "allthatcompany.com", "bizwiki.co.kr", "bizwiki.co", "foxsoft.kr",
    "kjuso.kr", "ggilbo.com", "hankooki.com", "cnews.co.kr", "boannews.com",
    "joongdo.co.kr", "slist.kr", "bntnews.co.kr", "energydaily.co.kr", "koswec.or.kr", 
    "114.co.kr", "kca.kr", "webify.kr", "danawa.com"
]

import sys
def safe_print(msg):
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        try:
            print(msg.encode(sys.stdout.encoding or 'utf-8', errors='replace').decode(sys.stdout.encoding or 'utf-8'), flush=True)
        except Exception:
            print(msg.encode('ascii', errors='replace').decode('ascii'), flush=True)

# Basic email regex
EMAIL_REGEX = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b')

def clean_emails(emails_set):
    cleaned = set()
    for email in emails_set:
        email_clean = email.strip().lower()
        if not email_clean or email_clean.endswith(".js"):
            continue
        domain = email_clean.split("@")[-1]
        if not any(junk in domain for junk in JUNK_DOMAINS):
            cleaned.add(email_clean)
    return cleaned

async def find_website_playwright(page, company_name):
    # Clean up corporate suffixes like (주) or ㈜ or 주식회사 for cleaner searches
    clean_name = company_name.replace("㈜", "").replace("(주)", "").replace("주식회사", "").strip()
    query = f"{clean_name} 홈페이지"
    search_url = f"https://search.naver.com/search.naver?query={urllib.parse.quote(query)}"
    
    try:
        safe_print(f"  [*] Searching Naver for: {clean_name}...")
        await page.goto(search_url, timeout=20000, wait_until="commit")
        await page.wait_for_timeout(2000)
        
        # Extract links
        links = await page.locator('a').all()
        candidate_urls = []
        for link in links:
            href = await link.get_attribute("href")
            if href and href.startswith(("http://", "https://")):
                href_lower = href.lower()
                if not any(domain in href_lower for domain in EXCLUDE_DOMAINS):
                    clean_url = href.split('?')[0].split('#')[0]
                    priority = 0
                    if clean_url.endswith((".co.kr", ".kr", ".com", ".net", ".org")):
                        priority = 2
                    elif ".co.kr/" in clean_url or ".kr/" in clean_url or ".com/" in clean_url:
                        priority = 1
                    candidate_urls.append((priority, clean_url))
        
        if candidate_urls:
            candidate_urls.sort(key=lambda x: x[0], reverse=True)
            return candidate_urls[0][1]
    except Exception as e:
        safe_print(f"  [-] Search failed for {clean_name}: {e}")
    return ""

async def crawl_site_for_emails(page, url):
    if not url:
        return []
    
    emails = set()
    try:
        safe_print(f"  [*] Navigating to Homepage: {url}...")
        await page.goto(url, timeout=20000, wait_until="commit")
        await page.wait_for_timeout(2000)
        
        # 1. Check body text
        body_text = await page.locator("body").inner_text()
        for email in EMAIL_REGEX.findall(body_text):
            emails.add(email.lower())
            
        # 2. Check HTML content
        html_content = await page.content()
        for email in EMAIL_REGEX.findall(html_content):
            emails.add(email.lower())
            
        # 3. Check mailto links
        mailto_links = await page.locator('a[href^="mailto:"]').all()
        for link in mailto_links:
            href = await link.get_attribute("href")
            if href:
                email = href.replace("mailto:", "").split("?")[0].strip().lower()
                emails.add(email)
                
        # 4. Search for contact pages
        links = await page.locator('a[href]').all()
        subpage_urls = []
        for link in links:
            href = await link.get_attribute("href")
            text = await link.inner_text()
            text_lower = text.lower() if text else ""
            href_lower = href.lower() if href else ""
            
            if href and not href_lower.startswith(('mailto:', 'tel:', 'javascript:', '#')):
                full_url = urllib.parse.urljoin(url, href)
                if urllib.parse.urlparse(full_url).netloc == urllib.parse.urlparse(url).netloc:
                    if any(k in text_lower or k in href_lower for k in ['contact', 'about', 'intro', 'location', 'company', '회사', '소개', '오시는']):
                        subpage_urls.append(full_url.split('#')[0])
                        
        for sub_url in list(set(subpage_urls))[:2]:
            try:
                safe_print(f"  [*] Navigating to Subpage: {sub_url}...")
                await page.goto(sub_url, timeout=12000, wait_until="commit")
                await page.wait_for_timeout(1000)
                sub_text = await page.locator("body").inner_text()
                for email in EMAIL_REGEX.findall(sub_text):
                    emails.add(email.lower())
            except Exception:
                pass
    except Exception as e:
        safe_print(f"  [-] Failed to crawl {url}: {e}")
        
    # Apply filtering
    return list(clean_emails(emails))

async def main_async():
    input_file = os.path.join(ROOT_DIR, "data", "raw", "kosma_instructors.csv")
    output_file = os.path.join(ROOT_DIR, "data", "formatted", "kosma_instructors_with_emails.csv")
    output_xlsx = os.path.join(ROOT_DIR, "data", "formatted", "kosma_instructors_with_emails.xlsx")
    
    if not os.path.exists(input_file):
        safe_print(f"[-] Input file {input_file} not found!")
        return
        
    # Read rows
    with open(input_file, mode="r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames
        if "Website" not in fieldnames:
            fieldnames.append("Website")
        if "Email" not in fieldnames:
            fieldnames.append("Email")
            
    safe_print(f"[*] Launching Playwright Chromium for {len(rows)} KOSMA companies...")
    
    total = len(rows)
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800}
        )
        page = await context.new_page()
        await stealth_async(page)
        
        for idx, row in enumerate(rows, 1):
            name = row.get("Tên công ty / Company Name", "")
            safe_print(f"\n[{idx}/{total}] Processing KOSMA: {name}...")
            
            # 1. Search for website
            website = await find_website_playwright(page, name)
            if website:
                row["Website"] = website
                safe_print(f"  [+] Found Website: {website}")
                
                # 2. Crawl website for emails
                emails = await crawl_site_for_emails(page, website)
                if emails:
                    row["Email"] = ", ".join(emails)
                    safe_print(f"  [+] Found Valid Emails: {row['Email']}")
                else:
                    row["Email"] = ""
                    safe_print("  [-] No valid emails found on site.")
            else:
                row["Website"] = ""
                row["Email"] = ""
                safe_print("  [-] No official website found.")
                
            # Random delay
            await page.wait_for_timeout(1500)
            
            # Save progress incrementally to CSV every 5 companies
            if idx % 5 == 0 or idx == total:
                with open(output_file, mode="w", encoding="utf-8-sig", newline="") as f_prog:
                    writer = csv.DictWriter(f_prog, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)
                safe_print(f"  [Progress] Saved incremental progress to CSV (processed {idx}/{total})")
        
        await browser.close()
        
    # Write to final Excel
    safe_print(f"[*] Writing results to Excel: {output_xlsx}...")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "KOSMA Managers with Emails"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws.append(fieldnames)
        for col_num, header in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        for r_idx, row in enumerate(rows, 1):
            row_data = [row.get(fn, "") for fn in fieldnames]
            ws.append(row_data)
            row_num = r_idx + 1
            
            is_even = (r_idx % 2 == 0)
            for col_num in range(1, len(fieldnames) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_all
                if is_even:
                    cell.fill = fill_zebra
                
                if col_num in [1, 4, 5, 7]: # STT, Phone, Fax, Reg Date
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws.row_dimensions[1].height = 28
        for r in range(2, len(rows) + 2):
            ws.row_dimensions[r].height = 20
            
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(output_xlsx)
        safe_print(f"[+] Saved final Excel: {os.path.abspath(output_xlsx)}")
    except Exception as e:
        safe_print(f"[-] Error writing Excel: {e}")

if __name__ == "__main__":
    asyncio.run(main_async())
