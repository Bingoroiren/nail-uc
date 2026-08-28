import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
import requests
from requests.adapters import HTTPAdapter
from urllib3.util import Retry
import re
import concurrent.futures
import csv
import os
import time
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Configure session with high retry count and backoff for weak network
session = requests.Session()
retries = Retry(
    total=3,              # 3 low-level retries per request
    backoff_factor=1,
    status_forcelist=[500, 502, 503, 504],
    raise_on_status=False
)
adapter = HTTPAdapter(pool_connections=8, pool_maxsize=8, max_retries=retries)
session.mount('https://', adapter)
session.mount('http://', adapter)

session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5"
})

csv_lock = threading.Lock()
global_stt = 1  # Global counter for real-time STT

def fetch_url(url):
    try:
        response = session.get(url, timeout=30)
        if response.status_code == 200:
            return response.text
        else:
            print(f"[-] Status {response.status_code} for URL: {url}", flush=True)
    except Exception as e:
        print(f"[-] Network error for URL {url}: {e}", flush=True)
    return None

def parse_list_page(html):
    companies = []
    tbody_match = re.search(r'<tbody>([\s\S]*?)</tbody>', html)
    if not tbody_match:
        return companies
    
    tbody_content = tbody_match.group(1)
    tr_blocks = re.findall(r'<tr>([\s\S]*?)</tr>', tbody_content)
    for tr in tr_blocks:
        tds = re.findall(r'<td[^>]*>([\s\S]*?)</td>', tr)
        if len(tds) >= 5:
            region = re.sub(r'<[^>]+>', '', tds[0]).strip()
            
            # Extract name and custCode
            name_a_match = re.search(r'href="[^"]*?custCode=([^"&]+)[^"]*"[^>]*>([\s\S]*?)</a>', tds[1])
            if name_a_match:
                cust_code = name_a_match.group(1).strip()
                company_name = re.sub(r'<[^>]+>', '', name_a_match.group(2)).strip()
            else:
                cust_code = ""
                company_name = re.sub(r'<[^>]+>', '', tds[1]).strip()
                
            representative = re.sub(r'<[^>]+>', '', tds[2]).strip()
            business_type = re.sub(r'<[^>]+>', '', tds[3]).strip()
            phone = re.sub(r'<[^>]+>', '', tds[4]).strip()
            
            companies.append({
                "region": region,
                "company_name": company_name,
                "cust_code": cust_code,
                "representative": representative,
                "business_type": business_type,
                "phone": phone
            })
    return companies

def extract_field(html, header_name):
    pattern = rf'<th>{re.escape(header_name)}</th>\s*<td[^>]*>([\s\S]*?)</td>'
    match = re.search(pattern, html)
    if match:
        val = match.group(1)
        val = re.sub(r'<[^>]+>', '', val)
        return " ".join(val.split())
    return ""

def fetch_detail(company):
    if not company["cust_code"]:
        company["fax"] = ""
        company["address"] = ""
        company["membership_type"] = ""
        return company, True
        
    url = f"https://theksa.or.kr/site/main/johab/pCustMst?custCode={company['cust_code']}"
    html = fetch_url(url)
    if html:
        company["fax"] = extract_field(html, "FAX")
        company["address"] = extract_field(html, "주소")
        company["membership_type"] = extract_field(html, "소유종류")
        return company, True
    else:
        return company, False

def main():
    global global_stt
    print("[*] STEP 1: Fetching list pages from KSA...", flush=True)
    start_time = time.time()
    
    total_pages = 232
    all_companies = []
    
    # Store list of pages that need fetching
    pending_pages = list(range(1, total_pages + 1))
    
    loop_count = 1
    # Run up to 6 passes to handle dropping packets
    while pending_pages and loop_count <= 6:
        print(f"\n[*] List Page Fetch - Pass {loop_count} ({len(pending_pages)} pages remaining)...", flush=True)
        failed_pages = []
        
        for page in pending_pages:
            url = f"https://theksa.or.kr/site/main/johab/pCustMstList?cp={page}&listType=list"
            html = fetch_url(url)
            if html:
                page_companies = parse_list_page(html)
                all_companies.extend(page_companies)
                print(f"[+] Page {page}/{total_pages} (Pass {loop_count}): Found {len(page_companies)} companies (Total: {len(all_companies)})", flush=True)
            else:
                print(f"[-] Failed to fetch page {page} on Pass {loop_count}", flush=True)
                failed_pages.append(page)
            
            # 0.1s delay between sequential calls
            time.sleep(0.1)
            
        pending_pages = failed_pages
        loop_count += 1
        if pending_pages:
            print(f"[*] Sleeping 5 seconds before retrying {len(pending_pages)} failed list pages...", flush=True)
            time.sleep(5)
            
    print(f"\n[+] Finished fetching list pages. Found {len(all_companies)} companies in total.", flush=True)
    print(f"[+] Listing fetch took {time.time() - start_time:.2f} seconds.", flush=True)
    
    # Initialize the CSV file with headers for real-time streaming
    csv_file = os.path.join(ROOT_DIR, "data", "raw", "ksa_members.csv")
    xlsx_file = os.path.join(ROOT_DIR, "data", "raw", "ksa_members.xlsx")
    
    headers = [
        "STT / No.",
        "Khu vực / Region",
        "Tên công ty / Company Name",
        "Người đại diện / Representative",
        "Số điện thoại / Phone",
        "Số fax / Fax",
        "Địa chỉ / Address",
        "Loại hình vận tải / Transport Type",
        "Loại hình hội viên / Membership Type",
        "Mã khách hàng / Cust Code"
    ]
    
    print(f"\n[*] Initializing CSV file: {csv_file} for incremental writing...", flush=True)
    try:
        with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
    except Exception as e:
        print(f"[-] Error initializing CSV: {e}", flush=True)
        
    # STEP 2: Fetch detail pages for each company concurrently with retry passes
    print("\n[*] STEP 2: Fetching detail pages (addresses, fax numbers, etc.) concurrently...", flush=True)
    detail_start_time = time.time()
    
    # Keep track of companies that still need details
    pending_companies = list(all_companies)
    successful_details = {}  # cust_code -> company_with_details
    
    loop_count = 1
    while pending_companies and loop_count <= 6:
        print(f"\n[*] Detail Fetch - Pass {loop_count} ({len(pending_companies)} companies remaining)...", flush=True)
        
        results_in_pass = []
        # Max 4 workers under weak network is much safer to avoid timeouts
        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            future_to_company = {executor.submit(fetch_detail, comp): comp for comp in pending_companies}
            
            count = 0
            for future in concurrent.futures.as_completed(future_to_company):
                res, success = future.result()
                results_in_pass.append((res, success))
                count += 1
                
                # If successfully fetched, write to CSV in real-time
                if success:
                    successful_details[res["cust_code"]] = res
                    # Lock thread writing to avoid corruption
                    with csv_lock:
                        try:
                            with open(csv_file, "a", newline="", encoding="utf-8-sig") as f:
                                writer = csv.writer(f)
                                writer.writerow([
                                    global_stt,
                                    res.get("region", "").strip(),
                                    res.get("company_name", "").strip(),
                                    res.get("representative", "").strip(),
                                    res.get("phone", "").strip(),
                                    res.get("fax", "").strip(),
                                    res.get("address", "").strip(),
                                    res.get("business_type", "").strip(),
                                    res.get("membership_type", "").strip(),
                                    res.get("cust_code", "").strip()
                                ])
                            global_stt += 1
                        except Exception as csv_write_err:
                            print(f"[-] CSV incremental write error: {csv_write_err}", flush=True)
                
                if count % 50 == 0 or count == len(pending_companies):
                    print(f"[+] Progress in Pass {loop_count}: {count}/{len(pending_companies)} completed", flush=True)
                    
        # Verify success and separate failures
        failed_companies = []
        for comp, success in results_in_pass:
            if not success:
                failed_companies.append(comp)
                
        pending_companies = failed_companies
        loop_count += 1
        if pending_companies:
            print(f"[*] Sleeping 10 seconds before retrying {len(pending_companies)} failed detail pages...", flush=True)
            time.sleep(10)
            
    print(f"[+] Finished fetching detail pages in {time.time() - detail_start_time:.2f} seconds.", flush=True)
    
    # STEP 3: Write final formatted files (sorting them by original page order)
    print(f"\n[*] Re-writing final sorted files...", flush=True)
    ordered_results = []
    
    # Write final CSV (sorted)
    try:
        with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
            # Align details with the order of all_companies
            for idx, orig_comp in enumerate(all_companies, 1):
                code = orig_comp["cust_code"]
                detail = successful_details.get(code, orig_comp)
                writer.writerow([
                    idx,
                    detail.get("region", "").strip(),
                    detail.get("company_name", "").strip(),
                    detail.get("representative", "").strip(),
                    detail.get("phone", "").strip(),
                    detail.get("fax", "").strip(),
                    detail.get("address", "").strip(),
                    detail.get("business_type", "").strip(),
                    detail.get("membership_type", "").strip(),
                    detail.get("cust_code", "").strip()
                ])
                ordered_results.append(detail)
                
        print(f"[+] Saved final sorted CSV file: {os.path.abspath(csv_file)}", flush=True)
    except Exception as e:
        print(f"[-] Error writing final sorted CSV: {e}", flush=True)
        ordered_results = list(successful_details.values())
        
    # Write final Excel
    print(f"[*] Writing final Excel file: {xlsx_file}...", flush=True)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "KSA Members"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        for idx, item in enumerate(ordered_results, 1):
            row_data = [
                idx,
                item.get("region", "").strip(),
                item.get("company_name", "").strip(),
                item.get("representative", "").strip(),
                item.get("phone", "").strip(),
                item.get("fax", "").strip(),
                item.get("address", "").strip(),
                item.get("business_type", "").strip(),
                item.get("membership_type", "").strip(),
                item.get("cust_code", "").strip()
            ]
            ws.append(row_data)
            row_num = idx + 1
            
            is_even = (idx % 2 == 0)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_all
                if is_even:
                    cell.fill = fill_zebra
                
                if col_num in [1, 10]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in [2, 4, 5, 6, 8, 9]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws.row_dimensions[1].height = 28
        for r in range(2, len(ordered_results) + 2):
            ws.row_dimensions[r].height = 20
            
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(xlsx_file)
        print(f"[+] Saved final Excel file successfully: {os.path.abspath(xlsx_file)}", flush=True)
    except Exception as e:
        print(f"[-] Error writing Excel: {e}", flush=True)
        
    print(f"\n[+] Total process took {time.time() - start_time:.2f} seconds.", flush=True)

if __name__ == "__main__":
    main()
