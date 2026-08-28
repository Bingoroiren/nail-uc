# format_korean_emails.py
# Formats and scrapes emails for Naver Map agencies, producing standard cold-mail files.

import asyncio
import csv
import os
import re
import sys
import random
import json
import urllib.parse
from playwright.async_api import async_playwright
from playwright_stealth import stealth_async
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Set console output encoding to UTF-8
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

INPUT_CSV = os.path.join(ROOT_DIR, "data", "raw", "korean_agencies.csv")
OUTPUT_CSV = os.path.join(ROOT_DIR, "data", "formatted", "korean_agencies_formatted.csv")
OUTPUT_XLSX = os.path.join(ROOT_DIR, "data", "formatted", "korean_agencies_formatted.xlsx")

EMAIL_REGEX = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b')

JUNK_DOMAINS = [
    "facebook.com", "twitter.com", "instagram.com", "linkedin.com", "youtube.com",
    "naver.com", "daum.net", "kakao.com", "wix.com", "wixsite.com", "wordpress.com",
    "squarespace.com", "weebly.com", "godaddy", "example.com", "domain.com", "placeholder",
    "wixpress"
]

CATEGORY_TRANSLATIONS = {
    "파견,헤드헌팅": "Phái cử, Headhunting (Môi giới nhân sự)",
    "직업안내": "Giới thiệu việc làm",
    "인력공급,고용알선": "Cung ứng nhân lực, Môi giới lao động",
    "외국인근로자센터": "Trung tâm lao động nước ngoài"
}

def safe_print(msg):
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        try:
            print(msg.encode(sys.stdout.encoding or 'utf-8', errors='replace').decode(sys.stdout.encoding or 'utf-8'), flush=True)
        except Exception:
            print(msg.encode('ascii', errors='replace').decode('ascii'), flush=True)



def score_email(email):
    email = email.lower().strip()
    if "@" not in email:
        return 0
    username, domain = email.split("@", 1)
    
    if any(email.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.js', '.css']):
        return 0
        
    system_usernames = ['noreply', 'no-reply', 'donotreply', 'privacy', 'terms', 'cookies', 'gdpr', 'abuse', 'security', 'webmaster', 'sentry', 'admin']
    if username in system_usernames or any(x in username for x in ['no-reply', 'noreply', 'privacy']):
        return 0
        
    junk_domains = JUNK_DOMAINS + ['sentry.io', 'wix.com', 'wordpress', 'squarespace', 'weebly', 'godaddy', 'example.com', 'domain.com', 'placeholder', 'wixpress', 'wixsite']
    if any(jd in domain for jd in junk_domains) or 'sentry' in domain or 'wix' in domain:
        return 0
        
    public_domains = ['gmail.com', 'yahoo.com', 'outlook.com', 'hotmail.com', 'aol.com', 'mail.com', 'live.com', 'msn.com', 'icloud.com', 'naver.com', 'daum.net', 'hanmail.net']
    if domain in public_domains:
        return 4
        
    generic_business_usernames = ['info', 'hello', 'office', 'contact', 'enquiries', 'sales', 'jobs', 'careers', 'recruitment', 'hr', 'work', 'apply', 'team', 'welcome']
    if username in generic_business_usernames:
        return 8
        
    return 10

async def crawl_site_for_emails(page, url):
    if not url:
        return []
    emails = set()
    try:
        await page.goto(url, timeout=12000, wait_until="commit")
        await page.wait_for_timeout(1000)
        
        # Scan body text & content
        body_text = await page.locator("body").inner_text()
        for email in EMAIL_REGEX.findall(body_text):
            emails.add(email.lower())
            
        html_content = await page.content()
        for email in EMAIL_REGEX.findall(html_content):
            emails.add(email.lower())
            
        # Scan mailto links
        mailto_links = await page.locator('a[href^="mailto:"]').all()
        for link in mailto_links:
            href = await link.get_attribute("href")
            if href:
                email = href.replace("mailto:", "").split("?")[0].strip().lower()
                if EMAIL_REGEX.match(email):
                    emails.add(email)
                    
        # Check subpages if no emails found
        if not emails:
            links = await page.locator('a[href]').all()
            subpage_urls = []
            for link in links:
                href = await link.get_attribute("href")
                text = await link.inner_text()
                text_lower = text.lower() if text else ""
                href_lower = href.lower() if href else ""
                
                if href and not href_lower.startswith(('mailto:', 'tel:', 'javascript:', '#')):
                    full_url = urllib.parse.urljoin(url, href)
                    if urllib.parse.urlparse(full_url).netloc == urllib.parse.urlparse(url).netloc:
                        if any(k in text_lower or k in href_lower for k in ['contact', 'about', 'intro', 'location', 'company', 'office', 'reach', 'mail']):
                            subpage_urls.append(full_url.split('#')[0])
                            
            for sub_url in list(set(subpage_urls))[:2]:
                try:
                    await page.goto(sub_url, timeout=8000, wait_until="commit")
                    await page.wait_for_timeout(1000)
                    sub_text = await page.locator("body").inner_text()
                    for email in EMAIL_REGEX.findall(sub_text):
                        emails.add(email.lower())
                except Exception:
                    pass
    except Exception:
        pass
        
    scored_emails = []
    for email in emails:
        score = score_email(email)
        if score >= 4:
            scored_emails.append((score, email))
            
    scored_emails.sort(key=lambda x: x[0], reverse=True)
    return [email for score, email in scored_emails]

async def main():
    if not os.path.exists(INPUT_CSV):
        safe_print(f"[-] Input file not found: {INPUT_CSV}")
        return
        
    safe_print(f"[*] Processing and crawling emails from: {INPUT_CSV}")
    
    with open(INPUT_CSV, mode="r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        input_rows = list(reader)
        
    # Translate headers to standard cold-mail columns
    fieldnames = [
        "No.", "Cong ty", "Chuc danh", "Nguoi lien he", "SDT", "Lien He", "Email", 
        "Lien He mail", "Dia chi", "Luong", "Ngay dang", "Han tuyen", "Check gui", 
        "Last Subject", "Last Body HTML", "Trang thai Reply", "Lan Follow-up", 
        "Ngay Follow-up gan nhat", "Mailbox da dung", "Category"
    ]
    
    # Load already crawled emails if output CSV exists
    existing_emails = {}
    if os.path.exists(OUTPUT_CSV):
        try:
            with open(OUTPUT_CSV, mode="r", encoding="utf-8-sig") as f_ex:
                reader_ex = csv.DictReader(f_ex)
                for r in reader_ex:
                    c_name = r.get("Cong ty", "")
                    if c_name:
                        existing_emails[c_name] = r.get("Email", "")
            safe_print(f"[+] Found existing output file. Loaded {len(existing_emails)} processed companies.")
        except Exception as e:
            safe_print(f"[-] Could not load existing output file: {e}")

    # Load progress JSON to keep track of processed companies
    PROGRESS_FILE_EMAILS = os.path.join(ROOT_DIR, "data", "progress", "scraping_progress_korean_emails.json")
    processed_companies = set()
    if os.path.exists(PROGRESS_FILE_EMAILS):
        try:
            with open(PROGRESS_FILE_EMAILS, "r", encoding="utf-8") as f_p:
                p_data = json.load(f_p)
                processed_companies = set(p_data.get("processed", []))
            safe_print(f"[+] Loaded {len(processed_companies)} already crawled companies from progress file.")
        except Exception as e:
            safe_print(f"[-] Error loading progress file: {e}")

    # Synchronize: any company with non-empty emails from OUTPUT_CSV is definitely processed
    for c_name, email in existing_emails.items():
        if email.strip():
            processed_companies.add(c_name)

    def save_processed_progress(processed_set):
        try:
            with open(PROGRESS_FILE_EMAILS, "w", encoding="utf-8") as f_p:
                json.dump({"processed": list(processed_set)}, f_p, indent=4, ensure_ascii=False)
        except Exception as e:
            pass

    formatted_rows = []
    
    # Pre-populate empty cold mail template, keeping existing emails if present
    for idx, row in enumerate(input_rows):
        comp_name = row.get("Tên công ty / Company Name", "")
        category = row.get("Danh mục / Category", "")
        phone = row.get("Số điện thoại / Phone", "")
        address = row.get("Địa chỉ / Address", "")
        website = row.get("Website", "")
        place_link = row.get("Naver Place Link", "")
        
        # Prepend ' to phone if missing
        if phone and not phone.startswith("'"):
            phone = f"'{phone}"
            
        # Category Translation
        trans_cat = CATEGORY_TRANSLATIONS.get(category, category)
        
        # Fallback website to place link if no official website
        website_to_use = website if website else place_link
        
        email_val = existing_emails.get(comp_name, "")
        
        formatted_rows.append({
            "No.": str(idx + 1),
            "Cong ty": comp_name,
            "Chuc danh": "",
            "Nguoi lien he": "",
            "SDT": phone,
            "Lien He": website_to_use,
            "Email": email_val,
            "Lien He mail": "",
            "Dia chi": address,
            "Luong": "",
            "Ngay dang": "",
            "Han tuyen": "",
            "Check gui": "",
            "Last Subject": "",
            "Last Body HTML": "",
            "Trang thai Reply": "",
            "Lan Follow-up": "0",
            "Ngay Follow-up gan nhat": "",
            "Mailbox da dung": "",
            "Category": trans_cat
        })
        
    # Email Crawling Session
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800}
        )
        page = await context.new_page()
        await stealth_async(page)
        
        total = len(formatted_rows)
        for i, row in enumerate(formatted_rows):
            comp_name = row["Cong ty"]
            web_url = row["Lien He"]
            
            if comp_name in processed_companies:
                if i % 100 == 0 or i == total - 1:
                    safe_print(f"[{i+1}/{total}] Skipping already processed: '{comp_name}'...")
                continue
                
            # Only crawl actual corporate websites (ignore place.naver.com links)
            is_valid_site = web_url.startswith("http") and "place.naver.com" not in web_url
            
            if is_valid_site:
                safe_print(f"[{i+1}/{total}] Crawling emails for: '{comp_name}' -> {web_url}...")
                emails = await crawl_site_for_emails(page, web_url)
                if emails:
                    row["Email"] = ", ".join(emails)
                    safe_print(f"  [+] Found: {row['Email']}")
                else:
                    safe_print("  [-] No emails found.")
            else:
                safe_print(f"[{i+1}/{total}] Skipping search for: '{comp_name}' (No valid website)")
                
            # Mark as processed and save incremental progress
            processed_companies.add(comp_name)
            save_processed_progress(processed_companies)
            
            # Intermittent saving to CSV
            with open(OUTPUT_CSV, mode="w", encoding="utf-8-sig", newline="") as f_out:
                writer = csv.DictWriter(f_out, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(formatted_rows)
                
            await page.wait_for_timeout(random.uniform(500, 1500))
            
        await browser.close()
        

        
    # Final Sorting: Push rows with emails to the top
    safe_print("[*] Sorting output: Emails on top...")
    with open(OUTPUT_CSV, mode="r", encoding="utf-8-sig") as f_out:
        reader = csv.DictWriter(f_out, fieldnames=fieldnames)
        rows_to_sort = list(csv.DictReader(f_out)) if hasattr(reader, "fieldnames") else []
        
    # Read output again safely
    with open(OUTPUT_CSV, mode="r", encoding="utf-8-sig") as f_out:
        reader = csv.DictReader(f_out)
        rows_to_sort = list(reader)
        
    rows_with_email = []
    rows_without_email = []
    for r in rows_to_sort:
        if r.get("Email", "").strip():
            rows_with_email.append(r)
        else:
            rows_without_email.append(r)
            
    sorted_rows = rows_with_email + rows_without_email
    for idx, r in enumerate(sorted_rows, 1):
        r["No."] = str(idx)
        
    # Write sorted CSV
    with open(OUTPUT_CSV, mode="w", encoding="utf-8-sig", newline="") as f_out:
        writer = csv.DictWriter(f_out, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(sorted_rows)
        
    # Write sorted XLSX
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Korean Agencies"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws.append(fieldnames)
        for col_num, h_name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        for r_idx, r in enumerate(sorted_rows, 1):
            row_data = [r.get(fn, "") for fn in fieldnames]
            ws.append(row_data)
            row_num = r_idx + 1
            
            is_even = (r_idx % 2 == 0)
            for col_num in range(1, len(fieldnames) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_all
                if is_even:
                    cell.fill = fill_zebra
                
                if col_num in [1, 5, 17]: # No., SDT, Lan Follow-up
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws.row_dimensions[1].height = 28
        for r in range(2, len(sorted_rows) + 2):
            ws.row_dimensions[r].height = 20
            
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(OUTPUT_XLSX)
        safe_print(f"[SUCCESS] Cleaned, crawled, sorted, and saved standard files: {OUTPUT_CSV} & {OUTPUT_XLSX}")
    except Exception as e:
        safe_print(f"[-] Error writing Excel output: {e}")

if __name__ == "__main__":
    asyncio.run(main())
