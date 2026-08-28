import csv
import os
import shutil
import re
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Set console output encoding to UTF-8
if sys.platform.startswith('win') and hasattr(sys.stdout, 'buffer'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)

CANDIDATES = [
    os.path.join(ROOT_DIR, "data", "formatted", "broker_canada_with_emails.csv"),
    os.path.join(ROOT_DIR, "data", "raw", "broker_canada.csv")
]

INPUT_CSV = None
for candidate in CANDIDATES:
    if os.path.exists(candidate):
        INPUT_CSV = candidate
        break

if len(sys.argv) > 1:
    INPUT_CSV = sys.argv[1]

if INPUT_CSV:
    ext = os.path.splitext(INPUT_CSV)[1]
    base_name = os.path.basename(INPUT_CSV)
    BACKUP_CSV = os.path.join(ROOT_DIR, "data", "formatted", base_name.replace(ext, f"_backup{ext}"))
    FORMATTED_CSV = os.path.join(ROOT_DIR, "data", "formatted", base_name.replace(ext, f"_formatted{ext}"))
    FORMATTED_XLSX = os.path.join(ROOT_DIR, "data", "formatted", base_name.replace(ext, f"_formatted.xlsx"))
    FORMATTED_V2_CSV = os.path.join(ROOT_DIR, "data", "formatted", base_name.replace(ext, f"_formatted_v2{ext}"))
    FORMATTED_V2_XLSX = os.path.join(ROOT_DIR, "data", "formatted", base_name.replace(ext, f"_formatted_v2.xlsx"))
else:
    BACKUP_CSV = None
    FORMATTED_CSV = None
    FORMATTED_XLSX = None
    FORMATTED_V2_CSV = None
    FORMATTED_V2_XLSX = None

GENERIC_DOMAINS = {
    'gmail.com', 'yahoo.com', 'hotmail.com', 'outlook.com', 'live.com',
    'aol.com', 'icloud.com', 'mail.com', 'ymail.com', 'msn.com',
    'wix.com', 'squarespace.com', 'wordpress.com', 'wixpress.com'
}

BUSINESS_PREFIXES = {
    'info', 'hello', 'contact', 'office', 'admin', 'sales',
    'enquiries', 'manager', 'service', 'hr', 'recruit', 'jobs', 'careers',
    'hiring', 'employment', 'staffing', 'talent'
}

BAD_KEYWORDS = {
    'wix', 'support', 'no-reply', 'noreply', 'test', 'example', 'domain'
}

import urllib.parse

SOCIAL_DOMAINS = {
    'facebook.com', 'fb.com', 'fb.me', 'instagram.com', 'instagr.am',
    'twitter.com', 'x.com', 'linkedin.com', 'youtube.com', 'youtu.be',
    'tiktok.com', 'pinterest.com', 'pin.it', 'reddit.com', 'threads.net',
    'tumblr.com', 'flickr.com', 'snapchat.com', 't.me', 'telegram.org',
    'telegram.me', 'wa.me', 'whatsapp.com', 'line.me', 'viber.com',
    'zalo.me', 'wechat.com', 'google.com', 'goo.gl', 'google.co.uk',
    'google.ie', 'google.com.au', 'google.ca', 'google.com.hk',
    'maps.app.goo.gl', 'waze.com', 'bing.com', 'yahoo.com',
    'duckduckgo.com', 'yelp.com', 'yelp.ca', 'yelp.co.uk',
    'tripadvisor.com', 'tripadvisor.ca', 'yellowpages.ca', 'yellowpages.com',
    'whitepages.com', 'foursquare.com', 'trustpilot.com', 'wikipedia.org',
    'wikidata.org', 'apple.com', 'apps.apple.com', 'play.google.com',
    'wix.com', 'wixsite.com', 'wixpress.com', 'squarespace.com',
    'wordpress.com', 'weebly.com', 'site123.me', 'jimdosite.com',
    'godaddysites.com', 'myshopify.com', 'canva.site', 'linktr.ee', 'carrd.co'
}

# Category translations from English to Vietnamese
CATEGORY_TRANSLATIONS = {
    "Employment agency": "Công ty môi giới lao động",
    "Temp agency": "Công ty cung ứng lao động thời vụ",
    "Human resource consulting": "Tư vấn nhân sự",
    "Recruiter": "Tuyển dụng",
    "Employment consultant": "Tư vấn việc làm",
    "employment agency": "Công ty môi giới lao động",
    "temp agency": "Công ty cung ứng lao động thời vụ",
    "human resource consulting": "Tư vấn nhân sự",
    "recruiter": "Tuyển dụng",
    "employment consultant": "Tư vấn việc làm",
}

def get_field(r, *keys):
    for k in keys:
        val = r.get(k, '')
        if val is not None and str(val).strip():
            return str(val).strip()
    return ''

def guess_email_from_website(website_url):
    if not website_url or not isinstance(website_url, str):
        return ""
    url = website_url.strip()
    if not url:
        return ""
    if not url.startswith(('http://', 'https://')):
        url = 'http://' + url
    try:
        parsed = urllib.parse.urlparse(url)
        netloc = parsed.netloc.strip().lower()
        if not netloc:
            return ""
        if ':' in netloc:
            netloc = netloc.split(':')[0]
        netloc = re.sub(r'^www\d*\.', '', netloc)
        if not netloc or '.' not in netloc:
            return ""
        for social in SOCIAL_DOMAINS:
            if netloc == social or netloc.endswith('.' + social):
                return ""
        if re.match(r'^[a-z0-9][a-z0-9\.\-]*\.[a-z]{2,}$', netloc):
            if netloc.endswith('.') or re.match(r'^\d+\.\d+\.\d+\.\d+$', netloc):
                return ""
            return f"info@{netloc}"
        return ""
    except Exception:
        return ""

def clean_and_score_email(email_str):
    if not email_str:
        return ""

    discard_domains = {
        'example.com', 'example.org', 'example.net', 'yourdomain.com',
        'email.com', 'domain.com', 'website.com', 'company.com',
        'sentry.io', 'git.com', 'github.com', 'test.com', 'g.co'
    }

    emails = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b', email_str)
    if not emails:
        return ""

    scored = []
    for em in emails:
        em_lower = em.strip().lower()

        if any(kw in em_lower for kw in BAD_KEYWORDS):
            continue

        parts = em_lower.split('@')
        if len(parts) != 2:
            continue

        username, domain = parts
        if domain in discard_domains:
            continue

        score = 10
        if domain in GENERIC_DOMAINS:
            score -= 3

        if any(username == pref for pref in BUSINESS_PREFIXES):
            score += 5
        elif any(pref in username for pref in BUSINESS_PREFIXES):
            score += 2

        scored.append((score, em_lower))

    if not scored:
        return ""

    scored.sort(key=lambda x: x[0], reverse=True)
    return scored[0][1]

def normalize_name(name):
    if not name:
        return ""
    n = name.lower()
    n = re.sub(r'[^a-z0-9\s]', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n

def normalize_phone(phone_str):
    if not phone_str:
        return ""
    p = re.sub(r'\D', '', phone_str)
    return p

def main():
    if not INPUT_CSV or not os.path.exists(INPUT_CSV):
        print(f"Error: Input file {INPUT_CSV} does not exist.")
        return

    print(f"[*] Backing up original CSV to {BACKUP_CSV}...")
    shutil.copy2(INPUT_CSV, BACKUP_CSV)

    rows = []
    with open(INPUT_CSV, mode='r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)

    print(f"[+] Loaded {len(rows)} rows.")

    # Filter closed
    active_rows = []
    closed_count = 0
    for r in rows:
        if r.get('Permanently_Closed', '').strip().lower() == 'yes':
            closed_count += 1
            continue
        active_rows.append(r)
    rows = active_rows
    if closed_count > 0:
        print(f"[*] Filtered out {closed_count} permanently closed businesses.")

    # Step 1: Normalize emails
    for r in rows:
        original_email = get_field(r, 'Email', 'Category', 'Lien He mail', 'email', 'Email lien he')
        r['Best_Email'] = clean_and_score_email(original_email)
        if not r['Best_Email']:
            website = r.get('Website') or r.get('Lien He') or r.get('URL') or r.get('website') or ''
            r['Best_Email'] = guess_email_from_website(website)

    # Step 2: Deduplicate by name
    grouped_rows = {}
    for r in rows:
        norm_name = normalize_name(get_field(r, 'Name', 'Cong ty', 'name', 'Ten cong ty'))
        if not norm_name:
            continue
        if norm_name not in grouped_rows:
            grouped_rows[norm_name] = []
        grouped_rows[norm_name].append(r)

    deduplicated_rows = []
    for norm_name, r_list in grouped_rows.items():
        def sort_rep(r):
            has_email = 1 if r['Best_Email'] else 0
            has_web = 1 if get_field(r, 'Website', 'Lien He', 'URL', 'website').strip() else 0
            try:
                reviews = float(r.get('Reviews_Count') or 0)
            except ValueError:
                reviews = 0.0
            try:
                rating = float(r.get('Rating') or 0)
            except ValueError:
                rating = 0.0
            return (-has_email, -has_web, -reviews, -rating)

        r_list.sort(key=sort_rep)
        deduplicated_rows.append(r_list[0])

    print(f"[+] Deduplicated to {len(deduplicated_rows)} unique company names.")

    # Step 3: Separate by email
    seen_emails = set()
    seen_phones = set()
    with_email = []
    without_email = []

    for r in deduplicated_rows:
        email = r['Best_Email']
        if not email:
            continue
        clean_email = email.strip().lower()
        phone = get_field(r, 'Phone', 'SDT', 'phone', 'Dien thoai')
        clean_phone = normalize_phone(phone)

        if clean_email in seen_emails:
            continue
        if clean_phone and clean_phone in seen_phones:
            continue

        seen_emails.add(clean_email)
        if clean_phone:
            seen_phones.add(clean_phone)
        with_email.append(r)

    for r in deduplicated_rows:
        email = r['Best_Email']
        if email:
            continue
        phone = get_field(r, 'Phone', 'SDT', 'phone', 'Dien thoai')
        clean_phone = normalize_phone(phone)

        if clean_phone and clean_phone in seen_phones:
            continue
        if clean_phone:
            seen_phones.add(clean_phone)
        without_email.append(r)

    print(f"[+] Found {len(with_email)} rows with email, {len(without_email)} rows without email.")
    final_sorted_rows = with_email + without_email

    # Step 4: Reformat to template
    fieldnames = [
        "No.", "Cong ty", "Chuc danh", "Nguoi lien he", "SDT", "Lien He",
        "Email", "Lien He mail", "Dia chi", "Luong", "Ngay dang", "Han tuyen",
        "Check gui", "Last Subject", "Last Body HTML", "Trang thai Reply",
        "Lan Follow-up", "Ngay Follow-up gan nhat", "Mailbox da dung", "Category"
    ]

    output_rows = []
    for i, r in enumerate(final_sorted_rows, 1):
        phone = get_field(r, 'Phone', 'SDT', 'phone', 'Dien thoai').strip()
        if phone:
            raw_digits = normalize_phone(phone)
            phone = f"'{raw_digits}"
        else:
            phone = ""

        raw_category = get_field(r, 'Category', 'category', 'Nganh nghe').strip()
        vi_category = CATEGORY_TRANSLATIONS.get(raw_category) or CATEGORY_TRANSLATIONS.get(raw_category.lower()) or raw_category

        out_row = {
            "No.": i,
            "Cong ty": get_field(r, 'Name', 'Cong ty', 'name', 'Ten cong ty'),
            "Chuc danh": "",
            "Nguoi lien he": "",
            "SDT": phone,
            "Lien He": get_field(r, 'Website', 'Lien He', 'URL', 'website'),
            "Email": r['Best_Email'],
            "Lien He mail": "",
            "Dia chi": get_field(r, 'Address', 'Dia chi', 'address'),
            "Luong": "",
            "Ngay dang": "",
            "Han tuyen": "",
            "Check gui": "",
            "Last Subject": "",
            "Last Body HTML": "",
            "Trang thai Reply": "",
            "Lan Follow-up": 0,
            "Ngay Follow-up gan nhat": "",
            "Mailbox da dung": "",
            "Category": vi_category
        }
        output_rows.append(out_row)

    # Write CSV
    target_write_file = FORMATTED_CSV
    try:
        with open(FORMATTED_CSV, mode='w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(output_rows)
    except PermissionError:
        target_write_file = FORMATTED_V2_CSV
        print(f"[!] {FORMATTED_CSV} is locked. Writing to {FORMATTED_V2_CSV} instead...")
        with open(FORMATTED_V2_CSV, mode='w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(output_rows)

    # Write Excel
    target_xlsx = FORMATTED_XLSX
    try:
        write_excel(output_rows, fieldnames, FORMATTED_XLSX)
    except PermissionError:
        target_xlsx = FORMATTED_V2_XLSX
        print(f"[!] {FORMATTED_XLSX} is locked. Writing to {FORMATTED_V2_XLSX} instead...")
        write_excel(output_rows, fieldnames, FORMATTED_V2_XLSX)

    print(f"\n--- Summary ---")
    print(f"Total output rows: {len(output_rows)}")
    print(f"With email:        {len(with_email)}")
    print(f"Without email:     {len(without_email)}")
    print(f"Formatted CSV:     {target_write_file}")
    print(f"Formatted XLSX:    {target_xlsx}")

def write_excel(rows, fieldnames, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Canada Brokers"

    ws.views.sheetView[0].showGridLines = True

    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, bold=False, color="000000")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    ws.append(fieldnames)
    for col_idx in range(1, len(fieldnames) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, r in enumerate(rows, 2):
        row_data = [r[col] for col in fieldnames]
        ws.append(row_data)

        has_email = bool(r["Email"])
        row_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") if has_email else None

        for col_idx in range(1, len(fieldnames) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

            col_name = fieldnames[col_idx - 1]
            if col_name in ["No.", "SDT", "Luong", "Ngay dang", "Han tuyen", "Check gui"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

            if col_name == "SDT" and cell.value:
                cell.number_format = '@'

    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith("'"):
                val_str = val_str[1:]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(file_path)

if __name__ == "__main__":
    main()
