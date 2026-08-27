import re
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def parse_kofa_html(html):
    companies = []
    
    # We will search for both regular and special member tables
    # 1. Regular members (정회원)
    regular_match = re.search(r'<h3>정회원</h3>([\s\S]*?)</table>', html)
    if regular_match:
        tr_blocks = re.findall(r'<tr[^>]*>([\s\S]*?)</tr>', regular_match.group(1))
        # Skip header row
        for tr in tr_blocks:
            tds = re.findall(r'<td[^>]*>([\s\S]*?)</td>', tr)
            if len(tds) >= 4:
                companies.append({
                    "name": re.sub(r'<[^>]+>', '', tds[0]).strip(),
                    "business": re.sub(r'<[^>]+>', '', tds[1]).strip(),
                    "phone": re.sub(r'<[^>]+>', '', tds[2]).strip(),
                    "address": re.sub(r'<[^>]+>', '', tds[3]).strip(),
                    "type": "Regular Member / 정회원"
                })
                
    # 2. Special members (특별회원)
    special_match = re.search(r'<h3>특별회원</h3>([\s\S]*?)</table>', html)
    if special_match:
        tr_blocks = re.findall(r'<tr[^>]*>([\s\S]*?)</tr>', special_match.group(1))
        for tr in tr_blocks:
            tds = re.findall(r'<td[^>]*>([\s\S]*?)</td>', tr)
            if len(tds) >= 4:
                companies.append({
                    "name": re.sub(r'<[^>]+>', '', tds[0]).strip(),
                    "business": re.sub(r'<[^>]+>', '', tds[1]).strip(),
                    "phone": re.sub(r'<[^>]+>', '', tds[2]).strip(),
                    "address": re.sub(r'<[^>]+>', '', tds[3]).strip(),
                    "type": "Special Member / 특별회원"
                })
                
    return companies

def main():
    html_path = r"C:\Users\Admin\.gemini\antigravity-ide\brain\f1cd8f96-441a-435a-9265-9c17ea67161b\.system_generated\steps\345\content.md"
    
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()
        
    companies = parse_kofa_html(html)
    print(f"[+] Successfully parsed {len(companies)} KOFA member fishing companies")
    
    headers = [
        "STT / No.",
        "Tên công ty / Company Name",
        "Loại hình đánh bắt / Fishery Type",
        "Số điện thoại / Phone",
        "Địa chỉ / Address",
        "Loại hình hội viên / Membership Type"
    ]
    
    csv_file = "kofa_members.csv"
    xlsx_file = "kofa_members.xlsx"
    
    # 1. Write CSV
    print(f"[*] Writing to CSV: {csv_file}")
    try:
        with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for idx, item in enumerate(companies, 1):
                writer.writerow([
                    idx,
                    item["name"],
                    item["business"],
                    item["phone"],
                    item["address"],
                    item["type"]
                ])
        print(f"[+] Saved CSV file successfully: {os.path.abspath(csv_file)}")
    except Exception as e:
        print(f"[-] Error writing CSV: {e}")
        
    # 2. Write Excel
    print(f"[*] Writing to Excel: {xlsx_file}")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "KOFA Fishing Members"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        for idx, item in enumerate(companies, 1):
            row_data = [
                idx,
                item["name"],
                item["business"],
                item["phone"],
                item["address"],
                item["type"]
            ]
            ws.append(row_data)
            row_num = idx + 1
            
            is_even = (idx % 2 == 0)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_all
                if is_even:
                    cell.fill = fill_zebra
                
                if col_num in [1, 6]: # STT, Membership type
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws.row_dimensions[1].height = 28
        for r in range(2, len(companies) + 2):
            ws.row_dimensions[r].height = 20
            
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(xlsx_file)
        print(f"[+] Saved Excel file successfully: {os.path.abspath(xlsx_file)}")
    except Exception as e:
        print(f"[-] Error writing Excel: {e}")

if __name__ == "__main__":
    main()
